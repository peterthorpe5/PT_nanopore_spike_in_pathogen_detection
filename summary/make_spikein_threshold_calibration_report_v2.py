#!/usr/bin/env python3
"""
Build a threshold-calibration report for ONT spike-in benchmarking data.

This script is designed to sit alongside the existing spike-in summary and
replicate-resolved reports without modifying them. It focuses specifically on
empirical cut-off selection for each workflow, metric, and target label.

The report is intended to answer a practical question:

- Below what raw detection value does the signal mostly resemble background?
- Above what raw detection value does the signal start to look like a true hit?

Rather than forcing a single universal threshold across all tools, the script
builds method-specific threshold recommendations from the observed negative and
positive distributions.

Outputs
-------
All tables are written as tab-separated files.

1. threshold_distribution_summary.tsv
   Negative and positive distribution summaries for each method.
2. threshold_scan.tsv
   Performance statistics evaluated across a scan of candidate thresholds.
3. threshold_recommendations.tsv
   Recommended thresholds using several decision rules.
4. threshold_recommendation_long.tsv
   Long-format version of the recommendation table, one row per rule.
5. threshold_plot_manifest.tsv
   Manifest of generated plots.
6. threshold_calibration_report.xlsx
   Formatted Excel workbook.
7. threshold_calibration_report.html
   Standalone HTML report.
8. plots/
   Distribution and threshold-scan plots.
"""

from __future__ import annotations

import argparse
import html
import math
from pathlib import Path
from typing import Optional

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter




METRIC_DEFINITIONS = [
    {
        "term": "n_positive / n_negative",
        "definition": (
            "Number of positive observations and negative observations used for "
            "threshold calibration within a workflow, metric, and target. "
            "Positives are non-zero spike-ins on the selected spike axis. "
            "Negatives are zero-spike observations and, where present, shuffled controls."
        ),
        "why_it_matters": (
            "Small negative counts make the empirical cut-off less stable. "
            "Small positive counts make sensitivity and LOD estimates less stable."
        ),
    },
    {
        "term": "TP / FN / FP / TN",
        "definition": (
            "True positives, false negatives, false positives, and true negatives "
            "after applying a candidate threshold."
        ),
        "why_it_matters": (
            "These are the core counts used to calculate all downstream performance statistics."
        ),
    },
    {
        "term": "Sensitivity",
        "definition": (
            "TP / (TP + FN). The proportion of genuine positive observations "
            "that are called detected at a given threshold."
        ),
        "why_it_matters": (
            "High sensitivity means the method is less likely to miss real low-level signal."
        ),
    },
    {
        "term": "Specificity",
        "definition": (
            "TN / (TN + FP). The proportion of negative observations that remain "
            "below the threshold."
        ),
        "why_it_matters": (
            "High specificity means the method is good at rejecting background noise."
        ),
    },
    {
        "term": "Precision",
        "definition": (
            "TP / (TP + FP). Among all observations called positive, the fraction "
            "that are genuinely positive."
        ),
        "why_it_matters": (
            "This is useful when you care about how trustworthy a positive call is."
        ),
    },
    {
        "term": "False positive rate",
        "definition": (
            "FP / (FP + TN). The proportion of negative observations that cross the threshold."
        ),
        "why_it_matters": (
            "This is the most direct summary of how often background is misclassified as real signal."
        ),
    },
    {
        "term": "False negative rate",
        "definition": (
            "FN / (FN + TP). The proportion of genuine positives that remain below the threshold."
        ),
        "why_it_matters": (
            "This shows the cost of using an overly conservative cut-off."
        ),
    },
    {
        "term": "Accuracy",
        "definition": (
            "(TP + TN) / (TP + TN + FP + FN). Overall fraction of correct calls."
        ),
        "why_it_matters": (
            "Useful as a broad summary, but it can be misleading when positives and negatives are unbalanced."
        ),
    },
    {
        "term": "Balanced accuracy",
        "definition": (
            "(Sensitivity + Specificity) / 2. Gives equal weight to positive and negative performance."
        ),
        "why_it_matters": (
            "Often more informative than raw accuracy when class sizes differ."
        ),
    },
    {
        "term": "F1 score",
        "definition": (
            "2TP / (2TP + FP + FN). Harmonic mean of precision and sensitivity."
        ),
        "why_it_matters": (
            "Useful when you want a single measure that penalises both missed positives and false positives."
        ),
    },
    {
        "term": "Youden's J",
        "definition": (
            "Sensitivity + Specificity - 1. A single summary of separation between positives and negatives."
        ),
        "why_it_matters": (
            "The balanced recommendation in this report chooses the threshold that maximises this quantity."
        ),
    },
    {
        "term": "LOD50 / LOD95 / LOD100",
        "definition": (
            "Lowest spike level at which the mean detection rate across positive observations "
            "reaches at least 50%, 95%, or 100% under the candidate threshold."
        ),
        "why_it_matters": (
            "These values summarise how early a method becomes reliably detectable as spike amount rises."
        ),
    },
    {
        "term": "95% confidence intervals",
        "definition": (
            "Wilson score intervals around sensitivity, specificity, precision, and false-positive rate."
        ),
        "why_it_matters": (
            "They show the uncertainty around the observed estimate. Wide intervals mean the estimate is less stable."
        ),
    },
    {
        "term": "Negative percentiles",
        "definition": (
            "The p01, p05, p10, p90, p95, and p99 columns summarise the spread of the negative background distribution."
        ),
        "why_it_matters": (
            "High upper negative percentiles indicate that background occasionally produces large values, "
            "which pushes empirical thresholds upwards."
        ),
    },
    {
        "term": "Negative mean plus 2 SD / 3 SD",
        "definition": (
            "Reference thresholds derived from the negative mean plus two or three standard deviations."
        ),
        "why_it_matters": (
            "These are familiar heuristic cut-offs, but they assume the negative spread is reasonably well behaved."
        ),
    },
    {
        "term": "Negative max plus step",
        "definition": (
            "Threshold set just above the largest observed negative value. For integer-valued metrics this is usually max negative + 1."
        ),
        "why_it_matters": (
            "This gives a conservative empirical cut-off with zero observed false positives in the current dataset."
        ),
    },
    {
        "term": "Positive below negative max fraction",
        "definition": (
            "Fraction of genuine positive observations whose raw value is still at or below the largest observed negative value."
        ),
        "why_it_matters": (
            "High values mean the positive and negative distributions overlap strongly, making thresholding intrinsically difficult."
        ),
    },
]


RECOMMENDATION_RULE_DEFINITIONS = [
    {
        "rule_name": "conservative_zero_observed_fp",
        "summary": "Just above the largest observed negative value.",
        "interpretation": (
            "Best when your priority is to avoid false positives in the current benchmark. "
            "The trade-off is that sensitivity may drop, especially at low spike levels."
        ),
    },
    {
        "rule_name": "balanced_youden_j",
        "summary": "Threshold that maximises Youden's J.",
        "interpretation": (
            "Best when you want a balanced compromise between recovering real signal and rejecting noise."
        ),
    },
    {
        "rule_name": "target_fpr",
        "summary": "Best-performing threshold among those with observed false-positive rate at or below the requested target.",
        "interpretation": (
            "Useful when you can tolerate some false positives but want to place an explicit upper limit on them."
        ),
    },
    {
        "rule_name": "negative_mean_plus_2sd",
        "summary": "Reference threshold from negative mean plus two standard deviations.",
        "interpretation": (
            "Useful as a familiar rule-of-thumb comparator, but less trustworthy when the negative distribution is skewed or sparse."
        ),
    },
]


HTML_STYLE = """
<style>
  body {
    font-family: Arial, Helvetica, sans-serif;
    margin: 0;
    padding: 0;
    color: #1a1a1a;
    background: #f7f9fc;
  }
  .container {
    max-width: 1480px;
    margin: 0 auto;
    padding: 32px;
    background: #ffffff;
  }
  h1, h2, h3 {
    color: #1f4e79;
  }
  h1 {
    margin-top: 0;
    font-size: 30px;
  }
  .lede {
    font-size: 16px;
    line-height: 1.5;
  }
  .muted {
    color: #666666;
  }
  .note-box {
    border-left: 4px solid #1f4e79;
    background: #f4f8fc;
    padding: 12px 16px;
    margin: 18px 0;
  }
  .summary-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 14px;
    margin: 24px 0;
  }
  .summary-card {
    border: 1px solid #d6dfeb;
    border-radius: 10px;
    padding: 16px;
    background: #f8fbff;
  }
  .summary-label {
    font-size: 13px;
    color: #4a5d73;
    margin-bottom: 8px;
  }
  .summary-value {
    font-size: 28px;
    font-weight: bold;
    color: #1f4e79;
  }
  .report-section {
    margin-top: 34px;
  }
  .table-wrap {
    overflow-x: auto;
    border: 1px solid #dbe5f0;
    border-radius: 10px;
    margin-top: 14px;
    background: #ffffff;
  }
  .data-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 14px;
  }
  .data-table thead th {
    position: sticky;
    top: 0;
    background: #1f4e79;
    color: white;
    text-align: left;
    padding: 10px 8px;
    white-space: nowrap;
    z-index: 2;
  }
  .data-table td {
    border-bottom: 1px solid #e6edf5;
    padding: 8px;
    vertical-align: top;
    white-space: nowrap;
  }
  .data-table tbody tr:nth-child(even) {
    background: #fbfdff;
  }
  .plot-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(360px, 1fr));
    gap: 18px;
  }
  .plot-card {
    border: 1px solid #d6dfeb;
    border-radius: 10px;
    padding: 12px;
    margin: 0;
    background: #ffffff;
  }
  .plot-card img {
    width: 100%;
    height: auto;
    display: block;
    border-radius: 6px;
  }
  .plot-card figcaption {
    margin-top: 8px;
    font-size: 13px;
    color: #4a5d73;
  }
  details {
    margin-top: 18px;
  }
  summary {
    cursor: pointer;
    font-weight: 600;
    color: #1f4e79;
  }
  @media print {
    body {
      background: white;
    }
    .container {
      max-width: none;
      margin: 0;
      padding: 16px;
    }
    a {
      color: inherit;
      text-decoration: none;
    }
  }
</style>
"""


def parse_args() -> argparse.Namespace:
    """
    Parse command-line arguments.

    Returns
    -------
    argparse.Namespace
        Parsed arguments.
    """
    parser = argparse.ArgumentParser(
        description=(
            "Build a threshold-calibration report from combined_long.tsv."
        )
    )
    parser.add_argument(
        "--summary_dir",
        required=True,
        help="Directory containing combined_long.tsv.",
    )
    parser.add_argument(
        "--out_dir",
        required=False,
        default=None,
        help="Output directory for the threshold-calibration report.",
    )
    parser.add_argument(
        "--title",
        required=False,
        default="ONT spike-in threshold calibration report",
        help="Title shown in the HTML report.",
    )
    parser.add_argument(
        "--spike_axis",
        choices=["total", "per_genome"],
        default="total",
        help="Spike axis used for positive and negative definitions.",
    )
    parser.add_argument(
        "--min_detect_value",
        type=float,
        default=1.0,
        help="Minimum permitted threshold.",
    )
    parser.add_argument(
        "--target_fpr",
        type=float,
        default=0.05,
        help="Target false-positive rate used for one recommendation rule.",
    )
    parser.add_argument(
        "--sd_multiplier",
        type=float,
        default=2.0,
        help="Multiplier used when reporting mean-plus-SD reference thresholds.",
    )
    return parser.parse_args()


def get_first_existing_column(
    *,
    dataframe: pd.DataFrame,
    candidates: list[str],
) -> Optional[str]:
    """
    Return the first candidate column present in a dataframe.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Input dataframe.
    candidates : list[str]
        Candidate column names.

    Returns
    -------
    Optional[str]
        The first matching column name, or None.
    """
    for candidate in candidates:
        if candidate in dataframe.columns:
            return candidate
    return None


def normalise_combined_long(*, combined_long: pd.DataFrame) -> pd.DataFrame:
    """
    Normalise the combined long table to a stable schema.

    Parameters
    ----------
    combined_long : pd.DataFrame
        Raw combined long table.

    Returns
    -------
    pd.DataFrame
        Normalised dataframe.
    """
    dataframe = combined_long.copy()
    rename_map: dict[str, str] = {}

    workflow_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["workflow", "workflow_type"],
    )
    metric_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["metric", "metric_name"],
    )
    value_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["value", "metric_value"],
    )
    target_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["target_label"],
    )
    run_name_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["run_name"],
    )
    shuffled_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["is_shuffled_control_recomputed", "is_shuffled_control"],
    )

    if workflow_column is None or metric_column is None or value_column is None:
        raise ValueError(
            "combined_long.tsv is missing one or more required columns: "
            "workflow/workflow_type, metric/metric_name, value/metric_value."
        )

    if workflow_column != "workflow":
        rename_map[workflow_column] = "workflow"
    if metric_column != "metric":
        rename_map[metric_column] = "metric"
    if value_column != "value":
        rename_map[value_column] = "value"
    if target_column and target_column != "target_label":
        rename_map[target_column] = "target_label"
    if run_name_column and run_name_column != "run_name":
        rename_map[run_name_column] = "run_name"
    if shuffled_column and shuffled_column != "is_shuffled_control":
        rename_map[shuffled_column] = "is_shuffled_control"

    if rename_map:
        dataframe = dataframe.rename(columns=rename_map)

    if "target_label" not in dataframe.columns:
        dataframe["target_label"] = "single_target"
    if "run_name" not in dataframe.columns:
        dataframe["run_name"] = dataframe["workflow"].astype(str)
    if "replicate" not in dataframe.columns:
        dataframe["replicate"] = pd.NA
    if "is_shuffled_control" not in dataframe.columns:
        dataframe["is_shuffled_control"] = False

    spike_total_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["total_spike_n", "spike_n", "total_spiked_reads"],
    )
    spike_per_genome_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["spike_n_per_genome", "spike_n"],
    )

    if spike_total_column is None and spike_per_genome_column is None:
        raise ValueError(
            "combined_long.tsv is missing a spike column such as total_spike_n "
            "or spike_n_per_genome."
        )

    dataframe["total_spike_n"] = pd.to_numeric(
        dataframe.get(spike_total_column, pd.NA),
        errors="coerce",
    )
    dataframe["spike_n_per_genome"] = pd.to_numeric(
        dataframe.get(spike_per_genome_column, pd.NA),
        errors="coerce",
    )

    if dataframe["total_spike_n"].isna().all():
        dataframe["total_spike_n"] = dataframe["spike_n_per_genome"]
    if dataframe["spike_n_per_genome"].isna().all():
        dataframe["spike_n_per_genome"] = dataframe["total_spike_n"]

    dataframe["workflow"] = dataframe["workflow"].fillna("unknown").astype(str)
    dataframe["metric"] = dataframe["metric"].fillna("unknown_metric").astype(str)
    dataframe["target_label"] = (
        dataframe["target_label"]
        .fillna("single_target")
        .replace("", "single_target")
        .astype(str)
    )
    dataframe["run_name"] = dataframe["run_name"].fillna("unknown_run").astype(str)
    dataframe["value"] = pd.to_numeric(dataframe["value"], errors="coerce")
    dataframe["replicate"] = pd.to_numeric(dataframe["replicate"], errors="coerce")
    dataframe["is_shuffled_control"] = (
        dataframe["is_shuffled_control"]
        .fillna(False)
        .astype(str)
        .str.lower()
        .isin(["true", "1", "yes"])
    )

    dataframe = dataframe.dropna(subset=["value"]).copy()
    dataframe["is_positive_total"] = (
        (~dataframe["is_shuffled_control"]) & (dataframe["total_spike_n"] > 0)
    )
    dataframe["is_negative_total"] = (
        dataframe["is_shuffled_control"] | (dataframe["total_spike_n"] == 0)
    )
    dataframe["is_positive_per_genome"] = (
        (~dataframe["is_shuffled_control"]) & (dataframe["spike_n_per_genome"] > 0)
    )
    dataframe["is_negative_per_genome"] = (
        dataframe["is_shuffled_control"] | (dataframe["spike_n_per_genome"] == 0)
    )
    return dataframe


def safe_divide(*, numerator: float, denominator: float) -> float:
    """
    Divide safely and return NaN on zero denominator.

    Parameters
    ----------
    numerator : float
        Numerator.
    denominator : float
        Denominator.

    Returns
    -------
    float
        Quotient or NaN.
    """
    if denominator == 0:
        return math.nan
    return numerator / denominator


def format_scalar(*, value: object) -> str:
    """
    Format a scalar for HTML display.

    Parameters
    ----------
    value : object
        Value to format.

    Returns
    -------
    str
        Formatted string.
    """
    if pd.isna(value):
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return f"{int(value):,}"
        if abs(value) >= 1000:
            return f"{value:,.3f}"
        return f"{value:.4f}"
    return html.escape(str(value))


def dataframe_to_html_table(
    *,
    dataframe: pd.DataFrame,
    max_rows: Optional[int] = None,
) -> str:
    """
    Render a dataframe as an HTML table.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Table to render.
    max_rows : Optional[int]
        Optional row cap for display.

    Returns
    -------
    str
        HTML table.
    """
    if dataframe is None or dataframe.empty:
        return '<p class="muted">No rows available.</p>'

    display_df = dataframe.copy()
    if max_rows is not None and display_df.shape[0] > max_rows:
        display_df = display_df.head(max_rows).copy()

    headers = "".join(
        f"<th>{html.escape(str(column))}</th>" for column in display_df.columns
    )
    body_rows: list[str] = []
    for _, row in display_df.iterrows():
        cells = "".join(
            f"<td>{format_scalar(value=row[column])}</td>"
            for column in display_df.columns
        )
        body_rows.append(f"<tr>{cells}</tr>")

    table = (
        '<div class="table-wrap"><table class="data-table">'
        f"<thead><tr>{headers}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody>"
        "</table></div>"
    )
    if max_rows is not None and dataframe.shape[0] > max_rows:
        table += (
            f'<p class="muted">Showing first {max_rows:,} of '
            f"{dataframe.shape[0]:,} rows in the HTML view.</p>"
        )
    return table


def build_summary_cards(*, entries: list[tuple[str, object]]) -> str:
    """
    Build HTML summary cards.

    Parameters
    ----------
    entries : list[tuple[str, object]]
        Label/value pairs.

    Returns
    -------
    str
        HTML snippet.
    """
    cards = []
    for label, value in entries:
        cards.append(
            "<div class=\"summary-card\">"
            f"<div class=\"summary-label\">{html.escape(str(label))}</div>"
            f"<div class=\"summary-value\">{html.escape(str(value))}</div>"
            "</div>"
        )
    return f'<div class="summary-grid">{"".join(cards)}</div>'


def build_plot_gallery(
    *,
    plot_manifest: pd.DataFrame,
    plot_type: str,
    relative_plot_dir: str,
) -> str:
    """
    Build an HTML plot gallery for one plot type.

    Parameters
    ----------
    plot_manifest : pd.DataFrame
        Plot manifest table.
    plot_type : str
        Plot type to filter.
    relative_plot_dir : str
        Relative plot directory path.

    Returns
    -------
    str
        HTML snippet.
    """
    subset = plot_manifest.loc[plot_manifest["plot_type"] == plot_type].copy()
    if subset.empty:
        return '<p class="muted">No plots available.</p>'

    cards: list[str] = []
    for _, row in subset.iterrows():
        png_href = f"{relative_plot_dir}/{row['png_path']}"
        caption = html.escape(str(row.get("caption", "")))
        title = html.escape(
            f"{row['workflow']} | {row['target_label']} | {row['metric']}"
        )
        cards.append(
            "<figure class=\"plot-card\">"
            f"<a href=\"{png_href}\" target=\"_blank\">"
            f"<img src=\"{png_href}\" alt=\"{title}\"></a>"
            f"<figcaption><strong>{title}</strong><br>{caption}</figcaption>"
            "</figure>"
        )
    return f'<div class="plot-grid">{"".join(cards)}</div>'


def wilson_interval(
    *,
    successes: int,
    total: int,
    z_value: float = 1.96,
) -> tuple[float, float]:
    """
    Calculate a Wilson confidence interval for a proportion.

    Parameters
    ----------
    successes : int
        Number of successes.
    total : int
        Number of trials.
    z_value : float
        Z-score for the desired confidence level.

    Returns
    -------
    tuple[float, float]
        Lower and upper confidence bounds.
    """
    if total <= 0:
        return (math.nan, math.nan)

    proportion = successes / total
    denominator = 1 + (z_value ** 2) / total
    centre = proportion + (z_value ** 2) / (2 * total)
    margin = z_value * math.sqrt(
        (proportion * (1 - proportion) / total)
        + ((z_value ** 2) / (4 * (total ** 2)))
    )
    lower = (centre - margin) / denominator
    upper = (centre + margin) / denominator
    return (max(0.0, lower), min(1.0, upper))


def next_threshold_above_max_negative(*, negative_values: pd.Series) -> float:
    """
    Return a threshold just above the observed negative maximum.

    Parameters
    ----------
    negative_values : pd.Series
        Negative values.

    Returns
    -------
    float
        Threshold just above the maximum observed negative value.
    """
    cleaned = pd.to_numeric(negative_values, errors="coerce").dropna()
    if cleaned.empty:
        return 0.0

    max_negative = float(cleaned.max())
    if np.all(np.isclose(cleaned.to_numpy(), np.round(cleaned.to_numpy()))):
        return float(max_negative + 1.0)
    return float(np.nextafter(max_negative, math.inf))


def first_spike_meeting_rate(
    *,
    dataframe: pd.DataFrame,
    spike_column: str,
    min_rate: float,
) -> float:
    """
    Return the first spike level at which the detection rate reaches a minimum.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Detection-call dataframe restricted to positives.
    spike_column : str
        Spike column to summarise.
    min_rate : float
        Minimum detection rate.

    Returns
    -------
    float
        First spike value meeting the requested rate, or NaN.
    """
    if dataframe.empty:
        return math.nan

    grouped = (
        dataframe.groupby(spike_column, dropna=False)["detected"]
        .mean()
        .reset_index()
        .sort_values(by=spike_column)
    )
    hits = grouped.loc[grouped["detected"] >= min_rate, spike_column]
    if hits.empty:
        return math.nan
    return float(hits.iloc[0])


def summarise_value_distribution(*, values: pd.Series, prefix: str) -> dict[str, object]:
    """
    Summarise a numeric value distribution.

    Parameters
    ----------
    values : pd.Series
        Numeric values.
    prefix : str
        Output prefix.

    Returns
    -------
    dict[str, object]
        Summary fields.
    """
    cleaned = pd.to_numeric(values, errors="coerce").dropna()
    if cleaned.empty:
        return {
            f"{prefix}_n": 0,
            f"{prefix}_min": math.nan,
            f"{prefix}_p01": math.nan,
            f"{prefix}_p05": math.nan,
            f"{prefix}_p10": math.nan,
            f"{prefix}_median": math.nan,
            f"{prefix}_mean": math.nan,
            f"{prefix}_sd": math.nan,
            f"{prefix}_p90": math.nan,
            f"{prefix}_p95": math.nan,
            f"{prefix}_p99": math.nan,
            f"{prefix}_max": math.nan,
        }

    return {
        f"{prefix}_n": int(cleaned.shape[0]),
        f"{prefix}_min": float(cleaned.min()),
        f"{prefix}_p01": float(cleaned.quantile(0.01)),
        f"{prefix}_p05": float(cleaned.quantile(0.05)),
        f"{prefix}_p10": float(cleaned.quantile(0.10)),
        f"{prefix}_median": float(cleaned.median()),
        f"{prefix}_mean": float(cleaned.mean()),
        f"{prefix}_sd": float(cleaned.std(ddof=0)),
        f"{prefix}_p90": float(cleaned.quantile(0.90)),
        f"{prefix}_p95": float(cleaned.quantile(0.95)),
        f"{prefix}_p99": float(cleaned.quantile(0.99)),
        f"{prefix}_max": float(cleaned.max()),
    }


def build_distribution_summary(
    *,
    dataframe: pd.DataFrame,
    positive_column: str,
    negative_column: str,
) -> pd.DataFrame:
    """
    Build negative and positive distribution summaries for each method.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Normalised combined-long dataframe.
    positive_column : str
        Positive flag column.
    negative_column : str
        Negative flag column.

    Returns
    -------
    pd.DataFrame
        Distribution summary table.
    """
    rows: list[dict[str, object]] = []
    grouped = dataframe.groupby(["workflow", "metric", "target_label"], dropna=False)

    for (workflow, metric, target_label), group in grouped:
        positives = group.loc[group[positive_column], "value"]
        negatives = group.loc[group[negative_column], "value"]

        record = {
            "workflow": workflow,
            "metric": metric,
            "target_label": target_label,
            **summarise_value_distribution(values=negatives, prefix="negative"),
            **summarise_value_distribution(values=positives, prefix="positive"),
        }
        record["negative_mean_plus_2sd"] = (
            record["negative_mean"] + (2.0 * record["negative_sd"])
            if not pd.isna(record["negative_mean"]) and not pd.isna(record["negative_sd"])
            else math.nan
        )
        record["negative_mean_plus_3sd"] = (
            record["negative_mean"] + (3.0 * record["negative_sd"])
            if not pd.isna(record["negative_mean"]) and not pd.isna(record["negative_sd"])
            else math.nan
        )
        record["negative_max_plus_step"] = next_threshold_above_max_negative(
            negative_values=negatives
        )
        record["positive_below_negative_max_fraction"] = safe_divide(
            numerator=int((pd.to_numeric(positives, errors="coerce") <= record["negative_max"]).sum())
            if not pd.isna(record["negative_max"])
            else 0,
            denominator=int(pd.to_numeric(positives, errors="coerce").dropna().shape[0]),
        )
        rows.append(record)

    out_df = pd.DataFrame(rows)
    if out_df.empty:
        return out_df
    return out_df.sort_values(by=["workflow", "metric", "target_label"]).reset_index(drop=True)


def calculate_performance_at_threshold(
    *,
    group: pd.DataFrame,
    threshold_value: float,
    positive_column: str,
    negative_column: str,
    spike_column: str,
) -> dict[str, object]:
    """
    Calculate detection performance at a specific threshold.

    Parameters
    ----------
    group : pd.DataFrame
        Group-specific dataframe.
    threshold_value : float
        Threshold being evaluated.
    positive_column : str
        Positive flag column.
    negative_column : str
        Negative flag column.
    spike_column : str
        Spike column.

    Returns
    -------
    dict[str, object]
        Performance summary.
    """
    evaluated = group.copy()
    evaluated["detected"] = evaluated["value"] >= threshold_value
    positives = evaluated.loc[evaluated[positive_column]].copy()
    negatives = evaluated.loc[evaluated[negative_column]].copy()

    tp = int((positives["detected"] == True).sum())
    fn = int((positives["detected"] == False).sum())
    fp = int((negatives["detected"] == True).sum())
    tn = int((negatives["detected"] == False).sum())

    sensitivity = safe_divide(numerator=tp, denominator=tp + fn)
    specificity = safe_divide(numerator=tn, denominator=tn + fp)
    precision = safe_divide(numerator=tp, denominator=tp + fp)
    false_positive_rate = safe_divide(numerator=fp, denominator=fp + tn)
    false_negative_rate = safe_divide(numerator=fn, denominator=fn + tp)
    accuracy = safe_divide(numerator=tp + tn, denominator=tp + tn + fp + fn)
    balanced_accuracy = (
        (sensitivity + specificity) / 2
        if not math.isnan(sensitivity) and not math.isnan(specificity)
        else math.nan
    )
    f1_score = safe_divide(numerator=2 * tp, denominator=(2 * tp) + fp + fn)
    youden_j = (
        sensitivity + specificity - 1
        if not math.isnan(sensitivity) and not math.isnan(specificity)
        else math.nan
    )

    sens_low, sens_high = wilson_interval(successes=tp, total=tp + fn)
    spec_low, spec_high = wilson_interval(successes=tn, total=tn + fp)
    prec_low, prec_high = wilson_interval(successes=tp, total=tp + fp)
    fpr_low, fpr_high = wilson_interval(successes=fp, total=fp + tn)

    lod50 = first_spike_meeting_rate(
        dataframe=positives,
        spike_column=spike_column,
        min_rate=0.50,
    )
    lod95 = first_spike_meeting_rate(
        dataframe=positives,
        spike_column=spike_column,
        min_rate=0.95,
    )
    lod100 = first_spike_meeting_rate(
        dataframe=positives,
        spike_column=spike_column,
        min_rate=1.00,
    )

    return {
        "threshold_value": float(threshold_value),
        "tp": tp,
        "fn": fn,
        "fp": fp,
        "tn": tn,
        "n_positive": int(positives.shape[0]),
        "n_negative": int(negatives.shape[0]),
        "sensitivity": sensitivity,
        "sensitivity_ci_low": sens_low,
        "sensitivity_ci_high": sens_high,
        "specificity": specificity,
        "specificity_ci_low": spec_low,
        "specificity_ci_high": spec_high,
        "precision": precision,
        "precision_ci_low": prec_low,
        "precision_ci_high": prec_high,
        "false_positive_rate": false_positive_rate,
        "false_positive_rate_ci_low": fpr_low,
        "false_positive_rate_ci_high": fpr_high,
        "false_negative_rate": false_negative_rate,
        "accuracy": accuracy,
        "balanced_accuracy": balanced_accuracy,
        "f1_score": f1_score,
        "youden_j": youden_j,
        "lod50_spike_n": lod50,
        "lod95_spike_n": lod95,
        "lod100_spike_n": lod100,
    }


def build_threshold_scan(
    *,
    dataframe: pd.DataFrame,
    positive_column: str,
    negative_column: str,
    spike_column: str,
    min_detect_value: float,
    target_fpr: float,
    sd_multiplier: float,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Build threshold-scan and recommendation tables.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Normalised combined-long dataframe.
    positive_column : str
        Positive flag column.
    negative_column : str
        Negative flag column.
    spike_column : str
        Spike column used for LOD summaries.
    min_detect_value : float
        Minimum permitted threshold.
    target_fpr : float
        Target false-positive rate.
    sd_multiplier : float
        Multiplier for a reported mean-plus-SD reference threshold.

    Returns
    -------
    tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]
        Threshold-scan table, wide recommendation table, and long
        recommendation table.
    """
    scan_rows: list[dict[str, object]] = []
    recommendation_rows: list[dict[str, object]] = []
    recommendation_long_rows: list[dict[str, object]] = []

    grouped = dataframe.groupby(["workflow", "metric", "target_label"], dropna=False)
    for (workflow, metric, target_label), group in grouped:
        values = pd.to_numeric(group["value"], errors="coerce").dropna()
        negatives = pd.to_numeric(
            group.loc[group[negative_column], "value"],
            errors="coerce",
        ).dropna()

        if values.empty:
            continue

        candidate_thresholds = set(float(value) for value in values.tolist())
        candidate_thresholds.add(float(min_detect_value))
        if not negatives.empty:
            candidate_thresholds.add(float(negatives.mean()))
            candidate_thresholds.add(float(negatives.max()))
            candidate_thresholds.add(
                float(negatives.mean() + (sd_multiplier * negatives.std(ddof=0)))
            )
            candidate_thresholds.add(
                next_threshold_above_max_negative(negative_values=negatives)
            )

        threshold_values = sorted(
            value for value in candidate_thresholds if not math.isnan(value)
        )

        scan_subset_rows: list[dict[str, object]] = []
        for threshold_value in threshold_values:
            performance = calculate_performance_at_threshold(
                group=group,
                threshold_value=max(float(min_detect_value), float(threshold_value)),
                positive_column=positive_column,
                negative_column=negative_column,
                spike_column=spike_column,
            )
            record = {
                "workflow": workflow,
                "metric": metric,
                "target_label": target_label,
                **performance,
            }
            scan_rows.append(record)
            scan_subset_rows.append(record)

        scan_subset_df = pd.DataFrame(scan_subset_rows)
        if scan_subset_df.empty:
            continue

        scan_subset_df = scan_subset_df.sort_values(by="threshold_value").reset_index(drop=True)

        conservative_threshold = max(
            float(min_detect_value),
            next_threshold_above_max_negative(negative_values=negatives),
        )
        conservative_row = calculate_performance_at_threshold(
            group=group,
            threshold_value=conservative_threshold,
            positive_column=positive_column,
            negative_column=negative_column,
            spike_column=spike_column,
        )

        balanced_df = scan_subset_df.copy()
        balanced_df["_youden_sort"] = balanced_df["youden_j"].fillna(-math.inf)
        balanced_df["_spec_sort"] = balanced_df["specificity"].fillna(-math.inf)
        balanced_df["_sens_sort"] = balanced_df["sensitivity"].fillna(-math.inf)
        balanced_row = (
            balanced_df.sort_values(
                by=["_youden_sort", "_spec_sort", "_sens_sort", "threshold_value"],
                ascending=[False, False, False, True],
            )
            .iloc[0]
            .to_dict()
        )

        target_fpr_df = scan_subset_df.loc[
            scan_subset_df["false_positive_rate"].fillna(math.inf) <= float(target_fpr)
        ].copy()
        if target_fpr_df.empty:
            target_fpr_row = balanced_row
            target_fpr_rule_note = "No threshold met the requested false-positive-rate target; fell back to balanced threshold."
        else:
            target_fpr_row = (
                target_fpr_df.sort_values(
                    by=["sensitivity", "youden_j", "threshold_value"],
                    ascending=[False, False, True],
                )
                .iloc[0]
                .to_dict()
            )
            target_fpr_rule_note = "Selected from thresholds with observed false-positive rate at or below the requested target."

        mean_plus_2sd_threshold = max(
            float(min_detect_value),
            float(negatives.mean() + (2.0 * negatives.std(ddof=0)))
            if not negatives.empty
            else float(min_detect_value),
        )
        mean_plus_2sd_row = calculate_performance_at_threshold(
            group=group,
            threshold_value=mean_plus_2sd_threshold,
            positive_column=positive_column,
            negative_column=negative_column,
            spike_column=spike_column,
        )

        recommendations = [
            (
                "conservative_zero_observed_fp",
                conservative_row,
                "Threshold set just above the maximum observed negative value, so no observed negatives exceed it in the current data.",
            ),
            (
                "balanced_youden_j",
                balanced_row,
                "Threshold chosen to maximise Youden's J, balancing sensitivity and specificity.",
            ),
            (
                f"target_fpr_le_{target_fpr:g}",
                target_fpr_row,
                target_fpr_rule_note,
            ),
            (
                "negative_mean_plus_2sd",
                mean_plus_2sd_row,
                "Reference threshold based on the negative mean plus two standard deviations.",
            ),
        ]

        wide_record: dict[str, object] = {
            "workflow": workflow,
            "metric": metric,
            "target_label": target_label,
            "min_detect_value": float(min_detect_value),
            "target_fpr": float(target_fpr),
            "n_positive": int(group.loc[group[positive_column]].shape[0]),
            "n_negative": int(group.loc[group[negative_column]].shape[0]),
        }

        for rule_name, row_data, note in recommendations:
            for key, value in row_data.items():
                wide_record[f"{rule_name}__{key}"] = value
            wide_record[f"{rule_name}__rule_note"] = note
            recommendation_long_rows.append(
                {
                    "workflow": workflow,
                    "metric": metric,
                    "target_label": target_label,
                    "rule_name": rule_name,
                    "rule_note": note,
                    **row_data,
                }
            )

        recommendation_rows.append(wide_record)

    scan_df = pd.DataFrame(scan_rows)
    recommendation_df = pd.DataFrame(recommendation_rows)
    recommendation_long_df = pd.DataFrame(recommendation_long_rows)

    if not scan_df.empty:
        scan_df = scan_df.sort_values(
            by=["workflow", "metric", "target_label", "threshold_value"]
        ).reset_index(drop=True)
    if not recommendation_df.empty:
        recommendation_df = recommendation_df.sort_values(
            by=["workflow", "metric", "target_label"]
        ).reset_index(drop=True)
    if not recommendation_long_df.empty:
        recommendation_long_df = recommendation_long_df.sort_values(
            by=["workflow", "metric", "target_label", "rule_name"]
        ).reset_index(drop=True)

    return scan_df, recommendation_df, recommendation_long_df


def maybe_symlog_x_axis(*, axis_values: pd.Series, ax: plt.Axes) -> None:
    """
    Apply a sensible symlog x-axis when values span a wide range.

    Parameters
    ----------
    axis_values : pd.Series
        Axis values.
    ax : matplotlib.axes.Axes
        Plot axis.
    """
    cleaned = pd.to_numeric(axis_values, errors="coerce").dropna()
    if cleaned.empty:
        return
    positive_values = cleaned.loc[cleaned > 0]
    if positive_values.empty:
        return
    minimum = float(positive_values.min())
    maximum = float(positive_values.max())
    if minimum <= 0 or maximum / minimum < 50:
        return
    ax.set_xscale("symlog", linthresh=max(1.0, minimum))


def sanitise_filename(*, value: str) -> str:
    """
    Convert a string into a filesystem-safe filename stem.

    Parameters
    ----------
    value : str
        Input string.

    Returns
    -------
    str
        Safe filename stem.
    """
    safe = "".join(character if character.isalnum() else "_" for character in value)
    while "__" in safe:
        safe = safe.replace("__", "_")
    return safe.strip("_")


def generate_distribution_plot(
    *,
    group_df: pd.DataFrame,
    positive_column: str,
    negative_column: str,
    conservative_threshold: float,
    balanced_threshold: float,
    out_dir: Path,
    workflow: str,
    metric: str,
    target_label: str,
) -> Optional[dict[str, str]]:
    """
    Generate a positive-versus-negative distribution plot.

    Parameters
    ----------
    group_df : pd.DataFrame
        Group-specific dataframe.
    positive_column : str
        Positive flag column.
    negative_column : str
        Negative flag column.
    conservative_threshold : float
        Conservative threshold.
    balanced_threshold : float
        Balanced threshold.
    out_dir : Path
        Plot output directory.
    workflow : str
        Workflow name.
    metric : str
        Metric name.
    target_label : str
        Target label.

    Returns
    -------
    Optional[dict[str, str]]
        Plot manifest entry.
    """
    positives = pd.to_numeric(
        group_df.loc[group_df[positive_column], "value"],
        errors="coerce",
    ).dropna()
    negatives = pd.to_numeric(
        group_df.loc[group_df[negative_column], "value"],
        errors="coerce",
    ).dropna()
    if positives.empty and negatives.empty:
        return None

    all_values = pd.concat([positives, negatives], ignore_index=True)
    if all_values.empty:
        return None

    value_min = float(all_values.min())
    value_max = float(all_values.max())
    if math.isclose(value_min, value_max):
        bins = np.linspace(value_min - 0.5, value_max + 0.5, 10)
    else:
        bins = np.linspace(value_min, value_max, 30)

    figure, axis = plt.subplots(figsize=(9, 5.2))
    if not negatives.empty:
        axis.hist(
            negatives,
            bins=bins,
            histtype="step",
            linewidth=2,
            density=False,
            label=f"Negatives (n={negatives.shape[0]})",
        )
    if not positives.empty:
        axis.hist(
            positives,
            bins=bins,
            histtype="step",
            linewidth=2,
            density=False,
            label=f"Positives (n={positives.shape[0]})",
        )

    if not math.isnan(conservative_threshold):
        axis.axvline(
            x=conservative_threshold,
            linestyle=":",
            linewidth=1.8,
            label=f"Conservative = {conservative_threshold:.4g}",
        )
    if not math.isnan(balanced_threshold):
        axis.axvline(
            x=balanced_threshold,
            linestyle="--",
            linewidth=1.8,
            label=f"Balanced = {balanced_threshold:.4g}",
        )

    axis.set_xlabel(metric.replace("_", " "))
    axis.set_ylabel("Observation count")
    axis.set_title(f"{workflow}: {target_label} - value distributions")
    axis.grid(True, alpha=0.3)
    axis.legend(loc="best", fontsize=8)
    figure.tight_layout()

    stem = sanitise_filename(
        value=f"distribution__{workflow}__{target_label}__{metric}"
    )
    png_path = out_dir / f"{stem}.png"
    pdf_path = out_dir / f"{stem}.pdf"
    figure.savefig(fname=png_path, dpi=200)
    figure.savefig(fname=pdf_path)
    plt.close(fig=figure)

    return {
        "plot_type": "distribution_plot",
        "workflow": workflow,
        "metric": metric,
        "target_label": target_label,
        "png_path": png_path.name,
        "pdf_path": pdf_path.name,
        "caption": (
            "Step histograms of negative and positive observations, with the "
            "conservative and balanced thresholds marked as vertical lines."
        ),
    }


def generate_threshold_scan_plot(
    *,
    scan_df: pd.DataFrame,
    conservative_threshold: float,
    balanced_threshold: float,
    out_dir: Path,
    workflow: str,
    metric: str,
    target_label: str,
) -> Optional[dict[str, str]]:
    """
    Generate a threshold-scan performance plot.

    Parameters
    ----------
    scan_df : pd.DataFrame
        Threshold-scan subset for one method.
    conservative_threshold : float
        Conservative threshold.
    balanced_threshold : float
        Balanced threshold.
    out_dir : Path
        Plot output directory.
    workflow : str
        Workflow name.
    metric : str
        Metric name.
    target_label : str
        Target label.

    Returns
    -------
    Optional[dict[str, str]]
        Plot manifest entry.
    """
    if scan_df.empty:
        return None

    plot_df = scan_df.sort_values(by="threshold_value").copy()
    figure, axis = plt.subplots(figsize=(9, 5.2))
    axis.plot(
        plot_df["threshold_value"],
        plot_df["sensitivity"],
        marker="o",
        linewidth=1.6,
        label="Sensitivity",
    )
    axis.plot(
        plot_df["threshold_value"],
        plot_df["specificity"],
        marker="o",
        linewidth=1.6,
        label="Specificity",
    )
    axis.plot(
        plot_df["threshold_value"],
        plot_df["precision"],
        marker="o",
        linewidth=1.6,
        label="Precision",
    )
    axis.plot(
        plot_df["threshold_value"],
        plot_df["youden_j"],
        marker="o",
        linewidth=1.6,
        label="Youden J",
    )

    if not math.isnan(conservative_threshold):
        axis.axvline(
            x=conservative_threshold,
            linestyle=":",
            linewidth=1.8,
            label=f"Conservative = {conservative_threshold:.4g}",
        )
    if not math.isnan(balanced_threshold):
        axis.axvline(
            x=balanced_threshold,
            linestyle="--",
            linewidth=1.8,
            label=f"Balanced = {balanced_threshold:.4g}",
        )

    maybe_symlog_x_axis(axis_values=plot_df["threshold_value"], ax=axis)
    axis.set_ylim(-0.05, 1.05)
    axis.set_xlabel("Threshold value")
    axis.set_ylabel("Performance statistic")
    axis.set_title(f"{workflow}: {target_label} - threshold scan ({metric})")
    axis.grid(True, alpha=0.3)
    axis.legend(loc="best", fontsize=8)
    figure.tight_layout()

    stem = sanitise_filename(
        value=f"threshold_scan__{workflow}__{target_label}__{metric}"
    )
    png_path = out_dir / f"{stem}.png"
    pdf_path = out_dir / f"{stem}.pdf"
    figure.savefig(fname=png_path, dpi=200)
    figure.savefig(fname=pdf_path)
    plt.close(fig=figure)

    return {
        "plot_type": "threshold_scan_plot",
        "workflow": workflow,
        "metric": metric,
        "target_label": target_label,
        "png_path": png_path.name,
        "pdf_path": pdf_path.name,
        "caption": (
            "Sensitivity, specificity, precision, and Youden's J across the "
            "candidate threshold range."
        ),
    }


def format_excel_workbook(
    *,
    workbook_path: Path,
    percentage_columns: set[str],
    integer_columns: set[str],
) -> None:
    """
    Apply workbook styling similar to the existing summary output.

    Parameters
    ----------
    workbook_path : Path
        Workbook path.
    percentage_columns : set[str]
        Columns shown with four decimals.
    integer_columns : set[str]
        Columns shown as integers.
    """
    workbook = load_workbook(filename=workbook_path)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(vertical="center", wrap_text=True)

    for worksheet in workbook.worksheets:
        if worksheet.max_row == 0 or worksheet.max_column == 0:
            continue

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        headers = [cell.value for cell in worksheet[1]]
        for idx, header in enumerate(headers, start=1):
            if header is None:
                continue
            header_text = str(header)
            column_letter = get_column_letter(idx)
            max_len = len(header_text)
            for row in worksheet.iter_rows(
                min_row=2,
                min_col=idx,
                max_col=idx,
                values_only=False,
            ):
                cell = row[0]
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
                    if header_text in percentage_columns and isinstance(cell.value, (int, float)):
                        cell.number_format = "0.0000"
                    elif header_text in integer_columns and isinstance(cell.value, (int, float)):
                        cell.number_format = "0"
            worksheet.column_dimensions[column_letter].width = min(max_len + 2, 42)

    workbook.save(filename=workbook_path)


def main() -> None:
    """
    Run the threshold-calibration reporting workflow.
    """
    args = parse_args()

    summary_dir = Path(args.summary_dir).expanduser().resolve()
    if not summary_dir.exists():
        raise FileNotFoundError(f"Summary directory not found: {summary_dir}")

    out_dir = (
        Path(args.out_dir).expanduser().resolve()
        if args.out_dir
        else summary_dir / "threshold_calibration_report"
    )
    out_dir.mkdir(parents=True, exist_ok=True)
    plots_dir = out_dir / "plots"
    plots_dir.mkdir(parents=True, exist_ok=True)

    combined_long_path = summary_dir / "combined_long.tsv"
    if not combined_long_path.exists():
        raise FileNotFoundError(f"Required file not found: {combined_long_path}")

    combined_long = pd.read_csv(filepath_or_buffer=combined_long_path, sep="\t")
    combined_long = normalise_combined_long(combined_long=combined_long)

    if args.spike_axis == "per_genome":
        positive_column = "is_positive_per_genome"
        negative_column = "is_negative_per_genome"
        spike_column = "spike_n_per_genome"
    else:
        positive_column = "is_positive_total"
        negative_column = "is_negative_total"
        spike_column = "total_spike_n"

    distribution_summary_df = build_distribution_summary(
        dataframe=combined_long,
        positive_column=positive_column,
        negative_column=negative_column,
    )
    threshold_scan_df, recommendation_df, recommendation_long_df = build_threshold_scan(
        dataframe=combined_long,
        positive_column=positive_column,
        negative_column=negative_column,
        spike_column=spike_column,
        min_detect_value=args.min_detect_value,
        target_fpr=args.target_fpr,
        sd_multiplier=args.sd_multiplier,
    )

    plot_rows: list[dict[str, str]] = []
    grouped = combined_long.groupby(["workflow", "metric", "target_label"], dropna=False)
    for (workflow, metric, target_label), group in grouped:
        long_subset = recommendation_long_df.loc[
            (recommendation_long_df["workflow"] == workflow)
            & (recommendation_long_df["metric"] == metric)
            & (recommendation_long_df["target_label"] == target_label)
        ].copy()
        scan_subset = threshold_scan_df.loc[
            (threshold_scan_df["workflow"] == workflow)
            & (threshold_scan_df["metric"] == metric)
            & (threshold_scan_df["target_label"] == target_label)
        ].copy()

        conservative_threshold = math.nan
        balanced_threshold = math.nan
        if not long_subset.empty:
            conservative_match = long_subset.loc[
                long_subset["rule_name"] == "conservative_zero_observed_fp"
            ]
            balanced_match = long_subset.loc[
                long_subset["rule_name"] == "balanced_youden_j"
            ]
            if not conservative_match.empty:
                conservative_threshold = float(
                    conservative_match["threshold_value"].iloc[0]
                )
            if not balanced_match.empty:
                balanced_threshold = float(balanced_match["threshold_value"].iloc[0])

        distribution_plot = generate_distribution_plot(
            group_df=group,
            positive_column=positive_column,
            negative_column=negative_column,
            conservative_threshold=conservative_threshold,
            balanced_threshold=balanced_threshold,
            out_dir=plots_dir,
            workflow=str(workflow),
            metric=str(metric),
            target_label=str(target_label),
        )
        if distribution_plot is not None:
            plot_rows.append(distribution_plot)

        scan_plot = generate_threshold_scan_plot(
            scan_df=scan_subset,
            conservative_threshold=conservative_threshold,
            balanced_threshold=balanced_threshold,
            out_dir=plots_dir,
            workflow=str(workflow),
            metric=str(metric),
            target_label=str(target_label),
        )
        if scan_plot is not None:
            plot_rows.append(scan_plot)

    plot_manifest_df = pd.DataFrame(plot_rows)

    output_tables: dict[str, pd.DataFrame] = {
        "threshold_distribution_summary": distribution_summary_df,
        "threshold_scan": threshold_scan_df,
        "threshold_recommendations": recommendation_df,
        "threshold_recommendation_long": recommendation_long_df,
        "threshold_plot_manifest": plot_manifest_df,
    }
    for stem, dataframe in output_tables.items():
        dataframe.to_csv(path_or_buf=out_dir / f"{stem}.tsv", sep="\t", index=False)

    workbook_path = out_dir / "threshold_calibration_report.xlsx"
    with pd.ExcelWriter(path=workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "metric": [
                    "n_observations",
                    "n_methods",
                    "n_threshold_scan_rows",
                    "n_plots",
                    "min_detect_value",
                    "target_fpr",
                    "spike_axis",
                ],
                "value": [
                    int(combined_long.shape[0]),
                    int(distribution_summary_df.shape[0]),
                    int(threshold_scan_df.shape[0]),
                    int(plot_manifest_df.shape[0]),
                    float(args.min_detect_value),
                    float(args.target_fpr),
                    args.spike_axis,
                ],
            }
        ).to_excel(writer, sheet_name="executive_summary", index=False)
        distribution_summary_df.to_excel(writer, sheet_name="distribution_summary", index=False)
        recommendation_df.to_excel(writer, sheet_name="recommendations", index=False)
        recommendation_long_df.to_excel(writer, sheet_name="recommendation_long", index=False)
        threshold_scan_df.to_excel(writer, sheet_name="threshold_scan", index=False)
        plot_manifest_df.to_excel(writer, sheet_name="plot_manifest", index=False)

    percentage_columns = {
        "value",
        "min_detect_value",
        "target_fpr",
        "threshold_value",
        "negative_min",
        "negative_p01",
        "negative_p05",
        "negative_p10",
        "negative_median",
        "negative_mean",
        "negative_sd",
        "negative_p90",
        "negative_p95",
        "negative_p99",
        "negative_max",
        "positive_min",
        "positive_p01",
        "positive_p05",
        "positive_p10",
        "positive_median",
        "positive_mean",
        "positive_sd",
        "positive_p90",
        "positive_p95",
        "positive_p99",
        "positive_max",
        "negative_mean_plus_2sd",
        "negative_mean_plus_3sd",
        "negative_max_plus_step",
        "positive_below_negative_max_fraction",
        "sensitivity",
        "sensitivity_ci_low",
        "sensitivity_ci_high",
        "specificity",
        "specificity_ci_low",
        "specificity_ci_high",
        "precision",
        "precision_ci_low",
        "precision_ci_high",
        "false_positive_rate",
        "false_positive_rate_ci_low",
        "false_positive_rate_ci_high",
        "false_negative_rate",
        "accuracy",
        "balanced_accuracy",
        "f1_score",
        "youden_j",
        "lod50_spike_n",
        "lod95_spike_n",
        "lod100_spike_n",
    }
    integer_columns = {
        "n_observations",
        "n_methods",
        "n_threshold_scan_rows",
        "n_plots",
        "negative_n",
        "positive_n",
        "tp",
        "fn",
        "fp",
        "tn",
        "n_positive",
        "n_negative",
        "replicate",
    }
    format_excel_workbook(
        workbook_path=workbook_path,
        percentage_columns=percentage_columns,
        integer_columns=integer_columns,
    )

    cards_html = build_summary_cards(
        entries=[
            ("Observations", f"{combined_long.shape[0]:,}"),
            ("Methods", f"{distribution_summary_df.shape[0]:,}"),
            ("Threshold scan rows", f"{threshold_scan_df.shape[0]:,}"),
            ("Plots", f"{plot_manifest_df.shape[0]:,}"),
            ("Minimum threshold", args.min_detect_value),
            ("Target false-positive rate", args.target_fpr),
            ("Spike axis", args.spike_axis),
        ]
    )

    html_page = f"""<!DOCTYPE html>
<html lang=\"en\">
<head>
  <meta charset=\"utf-8\">
  <title>{html.escape(args.title)}</title>
  {HTML_STYLE}
</head>
<body>
  <div class=\"container\">
    <h1>{html.escape(args.title)}</h1>
    <p class=\"lede\">
      Threshold-calibration report for the ONT spike-in framework. This report
      is designed to help you decide what raw signal value is still compatible
      with background noise for a given workflow, metric, and target, and what
      value begins to behave more like a true positive signal.
    </p>

    <div class=\"note-box\">
      <strong>How to interpret this report:</strong>
      Thresholds are method-specific rather than universal. The conservative
      recommendation is set just above the largest observed negative value and
      therefore gives zero observed false positives in the current dataset.
      The balanced recommendation maximises Youden's J and so usually gives a
      more permissive threshold with better sensitivity. Values below the
      conservative threshold still may be real in some settings, but in this
      dataset they overlap more with background.
    </div>

    <section class=\"report-section\">
      <h2>Executive summary</h2>
      {cards_html}
    </section>

    <section class="report-section">
      <h2>How the recommendation rules work</h2>
      <p class="muted">
        These are the decision rules used in the recommendation table. They are
        shown side by side because no single thresholding rule is best for
        every method or every intended use case.
      </p>
      {dataframe_to_html_table(dataframe=pd.DataFrame(RECOMMENDATION_RULE_DEFINITIONS), max_rows=20)}
    </section>

    <section class="report-section">
      <h2>Recommended thresholds</h2>
      <p class="muted">
        The wide table below brings together the main candidate thresholds for
        each workflow, metric, and target. The long-form version in the TSV and
        Excel output is easier to filter if you want one row per rule.
      </p>
      {dataframe_to_html_table(dataframe=recommendation_df, max_rows=100)}
    </section>

    <section class="report-section">
      <h2>Definitions of metrics used in this report</h2>
      <p class="muted">
        This section explains the main quantities shown in the tables and plots.
        Confidence intervals in the threshold scan are Wilson score intervals,
        which behave more sensibly than simple normal approximations when counts
        are small.
      </p>
      {dataframe_to_html_table(dataframe=pd.DataFrame(METRIC_DEFINITIONS), max_rows=50)}
    </section>

    <section class="report-section">
      <h2>Distribution summary</h2>
      <p class="muted">
        These columns summarise how negative and positive values overlap. The
        most important fields for threshold setting are the negative maximum,
        negative high percentiles, negative mean plus two or three standard
        deviations, and the fraction of positives that still lie at or below
        the negative maximum.
      </p>
      {dataframe_to_html_table(dataframe=distribution_summary_df, max_rows=100)}
    </section>

    <section class=\"report-section\">
      <h2>Threshold scan plots</h2>
      {build_plot_gallery(plot_manifest=plot_manifest_df, plot_type='threshold_scan_plot', relative_plot_dir='plots')}
    </section>

    <section class=\"report-section\">
      <h2>Distribution plots</h2>
      {build_plot_gallery(plot_manifest=plot_manifest_df, plot_type='distribution_plot', relative_plot_dir='plots')}
    </section>

    <section class=\"report-section\">
      <h2>Threshold scan detail</h2>
      <p class=\"muted\">
        This detailed table shows the full performance scan across all tested
        threshold values. Use the TSV or Excel workbook for complete filtering
        and sorting.
      </p>
      {dataframe_to_html_table(dataframe=threshold_scan_df, max_rows=300)}
    </section>
  </div>
</body>
</html>
"""

    html_path = out_dir / "threshold_calibration_report.html"
    html_path.write_text(data=html_page, encoding="utf-8")


if __name__ == "__main__":
    main()
