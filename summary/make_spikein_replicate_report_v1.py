#!/usr/bin/env python3
"""
Build a replicate-resolved ONT spike-in benchmarking report.

This script creates an additional report alongside the existing summary
workflow without modifying any of the current scripts. It reads the
``combined_long.tsv`` table produced by the spike-in summariser and writes a
new HTML report, a formatted Excel workbook, tab-separated detail tables, and
replicate-resolved plots.

Key differences from the existing stakeholder-facing report
-----------------------------------------------------------
- Every replicate is retained as an individual observation.
- Raw-signal plots show one line per run and replicate rather than a median
  line across replicates.
- Detection calls are written for every observation.
- Threshold tables are separated clearly from first-detection tables.
- Run-level and replicate-level summaries are both reported.

Optional support is included for ``reported_taxa_long.tsv`` so that taxa-level
thresholding and off-target review can also be inspected in a replicate-aware
way.
"""

from __future__ import annotations

import argparse
import html
import math
import re
from pathlib import Path
from typing import Optional

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HTML_STYLE = """
<style>
  body {
    font-family: Arial, Helvetica, sans-serif;
    margin: 0;
    padding: 0;
    color: #1a1a1a;
    background: #f7f9fc;
  }
  .container {
    max-width: 1480px;
    margin: 0 auto;
    padding: 32px;
    background: #ffffff;
  }
  h1, h2, h3 {
    color: #1f4e79;
  }
  h1 {
    margin-top: 0;
    font-size: 30px;
  }
  .lede {
    font-size: 16px;
    line-height: 1.5;
  }
  .muted {
    color: #666666;
  }
  .note-box {
    border-left: 4px solid #1f4e79;
    background: #f4f8fc;
    padding: 12px 16px;
    margin: 18px 0;
  }
  .summary-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 14px;
    margin: 24px 0;
  }
  .summary-card {
    border: 1px solid #d6dfeb;
    border-radius: 10px;
    padding: 16px;
    background: #f8fbff;
  }
  .summary-label {
    font-size: 13px;
    color: #4a5d73;
    margin-bottom: 8px;
  }
  .summary-value {
    font-size: 28px;
    font-weight: bold;
    color: #1f4e79;
  }
  .report-section {
    margin-top: 34px;
  }
  .table-wrap {
    overflow-x: auto;
    border: 1px solid #dbe5f0;
    border-radius: 10px;
    margin-top: 14px;
    background: #ffffff;
  }
  .data-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 14px;
  }
  .data-table thead th {
    position: sticky;
    top: 0;
    background: #1f4e79;
    color: white;
    text-align: left;
    padding: 10px 8px;
    white-space: nowrap;
    z-index: 2;
  }
  .data-table td {
    border-bottom: 1px solid #e6edf5;
    padding: 8px;
    vertical-align: top;
    white-space: nowrap;
  }
  .data-table tbody tr:nth-child(even) {
    background: #fbfdff;
  }
  .plot-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(360px, 1fr));
    gap: 18px;
  }
  .plot-card {
    border: 1px solid #d6dfeb;
    border-radius: 10px;
    padding: 12px;
    margin: 0;
    background: #ffffff;
  }
  .plot-card img {
    width: 100%;
    height: auto;
    display: block;
    border-radius: 6px;
  }
  .plot-card figcaption {
    margin-top: 8px;
    font-size: 13px;
    color: #4a5d73;
  }
  details {
    margin-top: 18px;
  }
  summary {
    cursor: pointer;
    font-weight: 600;
    color: #1f4e79;
  }
  @media print {
    body {
      background: white;
    }
    .container {
      max-width: none;
      margin: 0;
      padding: 16px;
    }
    a {
      color: inherit;
      text-decoration: none;
    }
  }
</style>
"""


def parse_args() -> argparse.Namespace:
    """
    Parse command-line arguments.

    Returns
    -------
    argparse.Namespace
        Parsed arguments.
    """
    parser = argparse.ArgumentParser(
        description=(
            "Build a replicate-resolved spike-in benchmarking report from "
            "combined_long.tsv and, optionally, reported_taxa_long.tsv."
        )
    )
    parser.add_argument(
        "--summary_dir",
        required=True,
        help="Directory containing combined_long.tsv and optional reported_taxa_long.tsv.",
    )
    parser.add_argument(
        "--out_dir",
        required=False,
        default=None,
        help="Output directory for the replicate-resolved report.",
    )
    parser.add_argument(
        "--title",
        required=False,
        default="ONT spike-in replicate-resolved report",
        help="Title shown in the HTML report.",
    )
    parser.add_argument(
        "--threshold_mode",
        choices=["fixed", "baseline_mean", "baseline_max", "baseline_mean_plus_sd"],
        default="fixed",
        help="Detection threshold rule.",
    )
    parser.add_argument(
        "--min_detect_value",
        type=float,
        default=1.0,
        help="Minimum raw metric value required for a positive detection.",
    )
    parser.add_argument(
        "--sd_multiplier",
        type=float,
        default=2.0,
        help=(
            "Standard-deviation multiplier used when threshold_mode is "
            "baseline_mean_plus_sd."
        ),
    )
    parser.add_argument(
        "--spike_axis",
        choices=["total", "per_genome"],
        default="total",
        help="Spike axis used for plots and first-detection summaries.",
    )
    return parser.parse_args()


def get_first_existing_column(
    *,
    dataframe: pd.DataFrame,
    candidates: list[str],
) -> Optional[str]:
    """
    Return the first candidate column present in a dataframe.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Input dataframe.
    candidates : list[str]
        Candidate column names.

    Returns
    -------
    Optional[str]
        The first matching column name, or None.
    """
    for candidate in candidates:
        if candidate in dataframe.columns:
            return candidate
    return None


def normalise_combined_long(*, combined_long: pd.DataFrame) -> pd.DataFrame:
    """
    Normalise the combined long table to a stable schema.

    Parameters
    ----------
    combined_long : pd.DataFrame
        Raw combined long table.

    Returns
    -------
    pd.DataFrame
        Normalised dataframe.
    """
    dataframe = combined_long.copy()
    rename_map: dict[str, str] = {}

    workflow_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["workflow", "workflow_type"],
    )
    metric_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["metric", "metric_name"],
    )
    value_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["value", "metric_value"],
    )
    target_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["target_label"],
    )
    run_name_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["run_name"],
    )
    run_dir_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["run_dir"],
    )
    shuffled_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["is_shuffled_control_recomputed", "is_shuffled_control"],
    )

    if workflow_column is None or metric_column is None or value_column is None:
        raise ValueError(
            "combined_long.tsv is missing one or more required columns: "
            "workflow/workflow_type, metric/metric_name, value/metric_value."
        )

    if workflow_column != "workflow":
        rename_map[workflow_column] = "workflow"
    if metric_column != "metric":
        rename_map[metric_column] = "metric"
    if value_column != "value":
        rename_map[value_column] = "value"
    if target_column and target_column != "target_label":
        rename_map[target_column] = "target_label"
    if run_name_column and run_name_column != "run_name":
        rename_map[run_name_column] = "run_name"
    if run_dir_column and run_dir_column != "run_dir":
        rename_map[run_dir_column] = "run_dir"
    if shuffled_column and shuffled_column != "is_shuffled_control":
        rename_map[shuffled_column] = "is_shuffled_control"

    if rename_map:
        dataframe = dataframe.rename(columns=rename_map)

    if "target_label" not in dataframe.columns:
        dataframe["target_label"] = "single_target"
    if "run_name" not in dataframe.columns:
        dataframe["run_name"] = dataframe["workflow"].astype(str)
    if "run_dir" not in dataframe.columns:
        dataframe["run_dir"] = pd.NA
    if "replicate" not in dataframe.columns:
        dataframe["replicate"] = pd.NA
    if "is_shuffled_control" not in dataframe.columns:
        dataframe["is_shuffled_control"] = False

    spike_total_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["total_spike_n", "spike_n", "total_spiked_reads"],
    )
    spike_per_genome_column = get_first_existing_column(
        dataframe=dataframe,
        candidates=["spike_n_per_genome", "spike_n"],
    )

    if spike_total_column is None and spike_per_genome_column is None:
        raise ValueError(
            "combined_long.tsv is missing a spike column such as total_spike_n "
            "or spike_n_per_genome."
        )

    dataframe["total_spike_n"] = pd.to_numeric(
        dataframe.get(spike_total_column, pd.NA),
        errors="coerce",
    )
    dataframe["spike_n_per_genome"] = pd.to_numeric(
        dataframe.get(spike_per_genome_column, pd.NA),
        errors="coerce",
    )

    if dataframe["total_spike_n"].isna().all():
        dataframe["total_spike_n"] = dataframe["spike_n_per_genome"]
    if dataframe["spike_n_per_genome"].isna().all():
        dataframe["spike_n_per_genome"] = dataframe["total_spike_n"]

    dataframe["workflow"] = dataframe["workflow"].fillna("unknown").astype(str)
    dataframe["metric"] = dataframe["metric"].fillna("unknown_metric").astype(str)
    dataframe["target_label"] = (
        dataframe["target_label"].fillna("single_target").replace("", "single_target").astype(str)
    )
    dataframe["run_name"] = dataframe["run_name"].fillna("unknown_run").astype(str)
    dataframe["run_dir"] = dataframe["run_dir"].fillna("").astype(str)
    dataframe["value"] = pd.to_numeric(dataframe["value"], errors="coerce")
    dataframe["replicate"] = pd.to_numeric(dataframe["replicate"], errors="coerce")
    dataframe["is_shuffled_control"] = (
        dataframe["is_shuffled_control"]
        .fillna(False)
        .astype(str)
        .str.lower()
        .isin(["true", "1", "yes"])
    )

    dataframe = dataframe.dropna(subset=["value"]).copy()
    dataframe["spike_n_plot"] = dataframe["total_spike_n"]
    dataframe["spike_n_plot_per_genome"] = dataframe["spike_n_per_genome"]
    dataframe["is_positive_total"] = (
        (~dataframe["is_shuffled_control"]) & (dataframe["total_spike_n"] > 0)
    )
    dataframe["is_negative_total"] = (
        dataframe["is_shuffled_control"] | (dataframe["total_spike_n"] == 0)
    )
    dataframe["is_positive_per_genome"] = (
        (~dataframe["is_shuffled_control"]) & (dataframe["spike_n_per_genome"] > 0)
    )
    dataframe["is_negative_per_genome"] = (
        dataframe["is_shuffled_control"] | (dataframe["spike_n_per_genome"] == 0)
    )
    return dataframe


def normalise_reported_taxa_long(*, dataframe: pd.DataFrame) -> pd.DataFrame:
    """
    Normalise the reported-taxa long table.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Raw reported-taxa table.

    Returns
    -------
    pd.DataFrame
        Normalised dataframe.
    """
    required = {"workflow_type", "run_name", "taxon_name", "metric_name", "metric_value"}
    missing = required - set(dataframe.columns)
    if missing:
        raise ValueError(
            "reported_taxa_long.tsv is missing required columns: "
            + ", ".join(sorted(missing))
        )

    out_df = dataframe.copy()
    out_df = out_df.rename(columns={"workflow_type": "workflow"})

    if "classifier" not in out_df.columns:
        out_df["classifier"] = "unknown_classifier"
    if "matched_target_labels" not in out_df.columns:
        out_df["matched_target_labels"] = ""
    if "replicate" not in out_df.columns:
        out_df["replicate"] = pd.NA
    if "spike_n_per_genome" not in out_df.columns:
        out_df["spike_n_per_genome"] = pd.NA
    if "total_spike_n" not in out_df.columns:
        out_df["total_spike_n"] = out_df["spike_n_per_genome"]
    if "n_genomes" not in out_df.columns:
        out_df["n_genomes"] = pd.NA
    if "is_shuffled_control" not in out_df.columns:
        out_df["is_shuffled_control"] = False

    out_df["workflow"] = out_df["workflow"].fillna("unknown").astype(str)
    out_df["run_name"] = out_df["run_name"].fillna("unknown_run").astype(str)
    out_df["classifier"] = out_df["classifier"].fillna("unknown_classifier").astype(str)
    out_df["taxon_name"] = out_df["taxon_name"].fillna("unknown_taxon").astype(str)
    out_df["metric_name"] = out_df["metric_name"].fillna("unknown_metric").astype(str)
    out_df["matched_target_labels"] = out_df["matched_target_labels"].fillna("").astype(str)
    out_df["metric_value"] = pd.to_numeric(out_df["metric_value"], errors="coerce")
    out_df["replicate"] = pd.to_numeric(out_df["replicate"], errors="coerce")
    out_df["spike_n_per_genome"] = pd.to_numeric(
        out_df["spike_n_per_genome"],
        errors="coerce",
    )
    out_df["total_spike_n"] = pd.to_numeric(out_df["total_spike_n"], errors="coerce")
    out_df["is_shuffled_control"] = (
        out_df["is_shuffled_control"]
        .fillna(False)
        .astype(str)
        .str.lower()
        .isin(["true", "1", "yes"])
    )
    out_df["spike_n"] = out_df["total_spike_n"].fillna(out_df["spike_n_per_genome"])
    out_df["is_positive"] = (~out_df["is_shuffled_control"]) & (out_df["spike_n"] > 0)
    out_df["is_negative"] = out_df["is_shuffled_control"] | (out_df["spike_n"] == 0)
    return out_df.dropna(subset=["metric_value"]).copy()


def choose_threshold(
    *,
    negative_values: pd.Series,
    threshold_mode: str,
    min_detect_value: float,
    sd_multiplier: float,
) -> float:
    """
    Choose a threshold from negative-control observations.

    Parameters
    ----------
    negative_values : pd.Series
        Negative-control values.
    threshold_mode : str
        Thresholding rule.
    min_detect_value : float
        Minimum permitted threshold.
    sd_multiplier : float
        Standard-deviation multiplier.

    Returns
    -------
    float
        Threshold value.
    """
    negative_values = pd.to_numeric(negative_values, errors="coerce").dropna()
    if negative_values.empty:
        return float(min_detect_value)
    if threshold_mode == "fixed":
        return float(min_detect_value)
    if threshold_mode == "baseline_mean":
        return float(max(min_detect_value, negative_values.mean()))
    if threshold_mode == "baseline_max":
        return float(max(min_detect_value, negative_values.max()))
    if threshold_mode == "baseline_mean_plus_sd":
        threshold = negative_values.mean() + (
            sd_multiplier * negative_values.std(ddof=0)
        )
        return float(max(min_detect_value, threshold))
    return float(min_detect_value)


def safe_divide(*, numerator: float, denominator: float) -> float:
    """
    Divide safely and return NaN on zero denominator.

    Parameters
    ----------
    numerator : float
        Numerator.
    denominator : float
        Denominator.

    Returns
    -------
    float
        Quotient or NaN.
    """
    if denominator == 0:
        return math.nan
    return numerator / denominator


def first_spike_meeting_rate(*, dataframe: pd.DataFrame, min_rate: float) -> float:
    """
    Return the first spike level at which the detection rate reaches a minimum.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Detection-call dataframe.
    min_rate : float
        Minimum rate.

    Returns
    -------
    float
        First spike level meeting the criterion.
    """
    if dataframe.empty:
        return math.nan

    grouped = (
        dataframe.groupby("spike_n", dropna=False)["detected"]
        .mean()
        .reset_index()
        .sort_values(by="spike_n")
    )
    detected = grouped.loc[grouped["detected"] >= min_rate, "spike_n"]
    if detected.empty:
        return math.nan
    return float(detected.iloc[0])


def wilson_interval(*, n_success: int, n_total: int, z: float = 1.959963984540054) -> tuple[float, float]:
    """
    Compute a Wilson score interval for a binomial proportion.

    Parameters
    ----------
    n_success : int
        Number of successes.
    n_total : int
        Number of observations.
    z : float
        Standard normal critical value.

    Returns
    -------
    tuple[float, float]
        Lower and upper confidence limits.
    """
    if n_total <= 0:
        return math.nan, math.nan

    phat = n_success / n_total
    z2 = z**2
    denominator = 1 + (z2 / n_total)
    centre = (phat + (z2 / (2 * n_total))) / denominator
    half_width = (
        z
        * math.sqrt(
            ((phat * (1 - phat)) / n_total) + (z2 / (4 * (n_total**2)))
        )
        / denominator
    )
    return max(0.0, centre - half_width), min(1.0, centre + half_width)


def compute_detection_calls(
    *,
    dataframe: pd.DataFrame,
    group_columns: list[str],
    value_column: str,
    positive_column: str,
    negative_column: str,
    spike_column: str,
    threshold_mode: str,
    min_detect_value: float,
    sd_multiplier: float,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Compute detection calls and thresholds.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Input dataframe.
    group_columns : list[str]
        Columns defining a method for thresholding.
    value_column : str
        Raw signal column.
    positive_column : str
        Boolean positive-observation column.
    negative_column : str
        Boolean negative-observation column.
    spike_column : str
        Spike column used for LOD summaries.
    threshold_mode : str
        Thresholding rule.
    min_detect_value : float
        Minimum threshold.
    sd_multiplier : float
        Standard-deviation multiplier.

    Returns
    -------
    tuple[pd.DataFrame, pd.DataFrame]
        Detection-call dataframe and threshold dataframe.
    """
    detection_rows: list[pd.DataFrame] = []
    threshold_rows: list[dict[str, object]] = []

    grouped = dataframe.groupby(group_columns, dropna=False)
    for keys, group in grouped:
        if not isinstance(keys, tuple):
            keys = (keys,)
        key_map = dict(zip(group_columns, keys))
        negatives = group.loc[group[negative_column], value_column]
        threshold_value = choose_threshold(
            negative_values=negatives,
            threshold_mode=threshold_mode,
            min_detect_value=min_detect_value,
            sd_multiplier=sd_multiplier,
        )
        negative_mean = float(negatives.mean()) if not negatives.dropna().empty else math.nan
        negative_max = float(negatives.max()) if not negatives.dropna().empty else math.nan
        negative_sd = float(negatives.std(ddof=0)) if not negatives.dropna().empty else math.nan
        n_negative = int(negatives.dropna().shape[0])

        group = group.copy()
        group["spike_n"] = pd.to_numeric(group[spike_column], errors="coerce")
        group["is_positive"] = group[positive_column].astype(bool)
        group["is_negative"] = group[negative_column].astype(bool)
        group["threshold_mode"] = threshold_mode
        group["threshold_value"] = threshold_value
        group["negative_mean"] = negative_mean
        group["negative_max"] = negative_max
        group["negative_sd"] = negative_sd
        group["n_negative_for_threshold"] = n_negative
        group["detected"] = pd.to_numeric(group[value_column], errors="coerce") >= threshold_value
        detection_rows.append(group)

        threshold_rows.append(
            {
                **key_map,
                "threshold_mode": threshold_mode,
                "threshold_value": threshold_value,
                "negative_mean": negative_mean,
                "negative_max": negative_max,
                "negative_sd": negative_sd,
                "n_negative_for_threshold": n_negative,
            }
        )

    detection_df = (
        pd.concat(detection_rows, ignore_index=True)
        if detection_rows
        else pd.DataFrame()
    )
    threshold_df = pd.DataFrame(threshold_rows)
    return detection_df, threshold_df


def summarise_performance(
    *,
    detection_df: pd.DataFrame,
    group_columns: list[str],
) -> pd.DataFrame:
    """
    Summarise confusion-matrix and LOD statistics.

    Parameters
    ----------
    detection_df : pd.DataFrame
        Detection-call dataframe.
    group_columns : list[str]
        Grouping columns.

    Returns
    -------
    pd.DataFrame
        Performance summary.
    """
    rows: list[dict[str, object]] = []
    grouped = detection_df.groupby(group_columns, dropna=False)

    for keys, group in grouped:
        if not isinstance(keys, tuple):
            keys = (keys,)
        key_map = dict(zip(group_columns, keys))

        positives = group.loc[group["is_positive"]].copy()
        negatives = group.loc[group["is_negative"]].copy()

        tp = int((positives["detected"] == True).sum())
        fn = int((positives["detected"] == False).sum())
        fp = int((negatives["detected"] == True).sum())
        tn = int((negatives["detected"] == False).sum())

        sensitivity = safe_divide(numerator=tp, denominator=tp + fn)
        specificity = safe_divide(numerator=tn, denominator=tn + fp)
        precision = safe_divide(numerator=tp, denominator=tp + fp)
        negative_predictive_value = safe_divide(numerator=tn, denominator=tn + fn)
        accuracy = safe_divide(numerator=tp + tn, denominator=tp + tn + fp + fn)
        balanced_accuracy = (
            (sensitivity + specificity) / 2
            if not math.isnan(sensitivity) and not math.isnan(specificity)
            else math.nan
        )
        false_positive_rate = safe_divide(numerator=fp, denominator=fp + tn)
        false_negative_rate = safe_divide(numerator=fn, denominator=fn + tp)
        f1_score = safe_divide(
            numerator=2 * tp,
            denominator=(2 * tp) + fp + fn,
        )
        youden_j = (
            sensitivity + specificity - 1
            if not math.isnan(sensitivity) and not math.isnan(specificity)
            else math.nan
        )
        lod50 = first_spike_meeting_rate(dataframe=positives, min_rate=0.50)
        lod95 = first_spike_meeting_rate(dataframe=positives, min_rate=0.95)
        lod100 = first_spike_meeting_rate(dataframe=positives, min_rate=1.00)

        rows.append(
            {
                **key_map,
                "n_positive": int(positives.shape[0]),
                "n_negative": int(negatives.shape[0]),
                "tp": tp,
                "fn": fn,
                "fp": fp,
                "tn": tn,
                "threshold_mode": group["threshold_mode"].iloc[0],
                "threshold_value": group["threshold_value"].iloc[0],
                "negative_mean": group["negative_mean"].iloc[0],
                "negative_max": group["negative_max"].iloc[0],
                "negative_sd": group["negative_sd"].iloc[0],
                "n_negative_for_threshold": group["n_negative_for_threshold"].iloc[0],
                "sensitivity": sensitivity,
                "specificity": specificity,
                "precision": precision,
                "negative_predictive_value": negative_predictive_value,
                "accuracy": accuracy,
                "balanced_accuracy": balanced_accuracy,
                "false_positive_rate": false_positive_rate,
                "false_negative_rate": false_negative_rate,
                "f1_score": f1_score,
                "youden_j": youden_j,
                "lod50_spike_n": lod50,
                "lod95_spike_n": lod95,
                "lod100_spike_n": lod100,
            }
        )

    return pd.DataFrame(rows)


def summarise_detection_by_spike(
    *,
    detection_df: pd.DataFrame,
    group_columns: list[str],
) -> pd.DataFrame:
    """
    Summarise detection rates by spike level.

    Parameters
    ----------
    detection_df : pd.DataFrame
        Detection-call dataframe.
    group_columns : list[str]
        Grouping columns excluding spike_n.

    Returns
    -------
    pd.DataFrame
        Detection-rate summary.
    """
    positive_df = detection_df.loc[detection_df["is_positive"]].copy()
    if positive_df.empty:
        return pd.DataFrame()

    grouped = (
        positive_df.groupby(group_columns + ["spike_n"], dropna=False)
        .agg(
            n_observations=("detected", "size"),
            n_detected=("detected", "sum"),
            detection_rate=("detected", "mean"),
        )
        .reset_index()
        .sort_values(by=group_columns + ["spike_n"])
        .reset_index(drop=True)
    )

    intervals = grouped.apply(
        lambda row: wilson_interval(
            n_success=int(row["n_detected"]),
            n_total=int(row["n_observations"]),
        ),
        axis=1,
    )
    grouped["detection_rate_ci_low"] = [interval[0] for interval in intervals]
    grouped["detection_rate_ci_high"] = [interval[1] for interval in intervals]
    return grouped


def summarise_signal_by_spike(
    *,
    dataframe: pd.DataFrame,
    value_column: str,
    group_columns: list[str],
    include_negatives: bool = True,
) -> pd.DataFrame:
    """
    Summarise raw signal by spike level.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Input dataframe.
    value_column : str
        Raw signal column.
    group_columns : list[str]
        Grouping columns excluding spike_n.
    include_negatives : bool
        Whether to include negative observations.

    Returns
    -------
    pd.DataFrame
        Signal summary.
    """
    signal_df = dataframe.copy()
    if not include_negatives:
        signal_df = signal_df.loc[signal_df["is_positive"]].copy()
    if signal_df.empty:
        return pd.DataFrame()

    grouped = (
        signal_df.groupby(group_columns + ["spike_n"], dropna=False)[value_column]
        .agg(["size", "mean", "median", "std", "min", "max"])
        .reset_index()
        .rename(columns={"size": "n_observations"})
        .sort_values(by=group_columns + ["spike_n"])
        .reset_index(drop=True)
    )
    return grouped


def summarise_replicate_first_detection(*, detection_df: pd.DataFrame) -> pd.DataFrame:
    """
    Summarise first detected spike per run and replicate.

    Parameters
    ----------
    detection_df : pd.DataFrame
        Detection-call dataframe.

    Returns
    -------
    pd.DataFrame
        Replicate-level summary.
    """
    rows: list[dict[str, object]] = []
    grouped = detection_df.groupby(
        ["workflow", "run_name", "replicate", "metric", "target_label"],
        dropna=False,
    )
    for keys, group in grouped:
        workflow, run_name, replicate, metric, target_label = keys
        positives = group.loc[group["is_positive"]].sort_values(by="spike_n")
        detected = positives.loc[positives["detected"]]
        first_detected = float(detected["spike_n"].iloc[0]) if not detected.empty else math.nan
        rows.append(
            {
                "workflow": workflow,
                "run_name": run_name,
                "replicate": replicate,
                "metric": metric,
                "target_label": target_label,
                "threshold_value": group["threshold_value"].iloc[0],
                "n_positive_spikes": int(positives.shape[0]),
                "n_detected_positive_spikes": int(detected.shape[0]),
                "first_detected_spike": first_detected,
                "max_signal": float(pd.to_numeric(group["value"], errors="coerce").max()),
            }
        )
    return pd.DataFrame(rows)


def summarise_thresholds_with_first_detection(
    *,
    threshold_df: pd.DataFrame,
    by_spike_df: pd.DataFrame,
    group_columns: list[str],
) -> pd.DataFrame:
    """
    Add first-detection spike summaries to a threshold table.

    Parameters
    ----------
    threshold_df : pd.DataFrame
        Threshold table.
    by_spike_df : pd.DataFrame
        Detection-by-spike summary.
    group_columns : list[str]
        Grouping columns shared by both tables.

    Returns
    -------
    pd.DataFrame
        Augmented threshold summary.
    """
    if threshold_df.empty:
        return threshold_df.copy()

    rows: list[dict[str, object]] = []
    grouped = by_spike_df.groupby(group_columns, dropna=False)
    first_detection_map: dict[tuple[object, ...], dict[str, float]] = {}

    for keys, group in grouped:
        if not isinstance(keys, tuple):
            keys = (keys,)
        ordered = group.sort_values(by="spike_n")
        first_detection_map[keys] = {
            "first_detected_spike_any": _first_rate_from_group(dataframe=ordered, min_rate=1e-12),
            "first_detected_spike_50pct": _first_rate_from_group(dataframe=ordered, min_rate=0.50),
            "first_detected_spike_95pct": _first_rate_from_group(dataframe=ordered, min_rate=0.95),
            "first_detected_spike_all": _first_rate_from_group(dataframe=ordered, min_rate=1.00),
        }

    for _, row in threshold_df.iterrows():
        key = tuple(row[column] for column in group_columns)
        extra = first_detection_map.get(
            key,
            {
                "first_detected_spike_any": math.nan,
                "first_detected_spike_50pct": math.nan,
                "first_detected_spike_95pct": math.nan,
                "first_detected_spike_all": math.nan,
            },
        )
        rows.append({**row.to_dict(), **extra})

    return pd.DataFrame(rows)


def _first_rate_from_group(*, dataframe: pd.DataFrame, min_rate: float) -> float:
    """
    Return the first spike from a by-spike summary meeting a minimum rate.

    Parameters
    ----------
    dataframe : pd.DataFrame
        By-spike summary for one group.
    min_rate : float
        Minimum detection rate.

    Returns
    -------
    float
        First spike level meeting the criterion.
    """
    detected = dataframe.loc[dataframe["detection_rate"] >= min_rate, "spike_n"]
    if detected.empty:
        return math.nan
    return float(detected.iloc[0])


def choose_reported_taxa_metric(*, dataframe: pd.DataFrame) -> pd.DataFrame:
    """
    Restrict the taxa table to the main clade-read metric when available.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Normalised reported-taxa table.

    Returns
    -------
    pd.DataFrame
        Filtered dataframe.
    """
    metric_mask = dataframe["metric_name"].astype(str).str.endswith(
        "_reported_clade_reads"
    )
    subset = dataframe.loc[metric_mask].copy()
    if subset.empty:
        subset = dataframe.copy()
    return subset


def summarise_reported_taxa_found(
    *,
    detection_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Summarise reported taxa found above threshold by run.

    Parameters
    ----------
    detection_df : pd.DataFrame
        Reported-taxa detection calls.

    Returns
    -------
    pd.DataFrame
        Run-level taxa summary.
    """
    if detection_df.empty:
        return pd.DataFrame()

    rows: list[dict[str, object]] = []
    grouped = detection_df.groupby(["workflow", "classifier", "run_name"], dropna=False)
    for keys, group in grouped:
        workflow, classifier, run_name = keys
        positives = group.loc[group["is_positive"] & group["detected"]].copy()
        taxa = sorted(pd.unique(positives["taxon_name"].astype(str)))
        rows.append(
            {
                "workflow": workflow,
                "classifier": classifier,
                "run_name": run_name,
                "n_taxa_found": len(taxa),
                "taxa_found": "; ".join(taxa),
            }
        )
    return pd.DataFrame(rows)


def maybe_symlog_x_axis(*, axis_values: pd.Series, ax: plt.Axes) -> None:
    """
    Apply a sensible symlog x-axis when spike levels span a wide range.

    Parameters
    ----------
    axis_values : pd.Series
        Spike values.
    ax : matplotlib.axes.Axes
        Plot axis.
    """
    positive = pd.to_numeric(axis_values, errors="coerce")
    positive = positive.loc[(positive > 0) & pd.notna(positive)]
    if positive.empty:
        return
    if positive.max() / max(positive.min(), 1) < 50:
        return
    ax.set_xscale("symlog", linthresh=max(1, float(positive.min())))


def sanitise_filename(*, value: str) -> str:
    """
    Convert a label into a filesystem-safe stem.

    Parameters
    ----------
    value : str
        Raw label.

    Returns
    -------
    str
        Safe stem.
    """
    cleaned = value.strip().replace(" ", "_")
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", cleaned)
    cleaned = re.sub(r"_+", "_", cleaned)
    return cleaned.strip("_") or "plot"


def make_trace_label(*, run_name: str, replicate: object, is_shuffled: bool) -> str:
    """
    Create a readable plot label for a run and replicate.

    Parameters
    ----------
    run_name : str
        Run name.
    replicate : object
        Replicate identifier.
    is_shuffled : bool
        Whether the trace is a shuffled control.

    Returns
    -------
    str
        Plot label.
    """
    if pd.notna(replicate):
        label = f"{run_name} | rep {int(replicate)}"
    else:
        label = run_name
    if is_shuffled:
        label = f"{label} (shuffled)"
    return label


def generate_raw_signal_plot(
    *,
    plot_df: pd.DataFrame,
    threshold_value: float,
    out_dir: Path,
    workflow: str,
    metric: str,
    target_label: str,
) -> Optional[dict[str, str]]:
    """
    Generate a raw-signal replicate trace plot.

    Parameters
    ----------
    plot_df : pd.DataFrame
        Subset for one workflow, metric, and target.
    threshold_value : float
        Detection threshold shown as a horizontal line.
    out_dir : Path
        Plot output directory.
    workflow : str
        Workflow name.
    metric : str
        Metric name.
    target_label : str
        Target label.

    Returns
    -------
    Optional[dict[str, str]]
        Plot manifest entry.
    """
    if plot_df.empty:
        return None

    figure, axis = plt.subplots(figsize=(8, 5))
    plot_df = plot_df.sort_values(by=["spike_n", "replicate", "run_name"], kind="stable")

    for keys, rep_df in plot_df.groupby(["run_name", "replicate"], dropna=False):
        run_name, replicate = keys
        rep_df = rep_df.sort_values(by="spike_n")
        label = make_trace_label(
            run_name=str(run_name),
            replicate=replicate,
            is_shuffled=bool(rep_df["is_shuffled_control"].iloc[0]),
        )
        linestyle = "--" if bool(rep_df["is_shuffled_control"].iloc[0]) else "-"
        axis.plot(
            rep_df["spike_n"],
            rep_df["value"],
            marker="o",
            linestyle=linestyle,
            linewidth=1.5,
            label=label,
        )

    if not math.isnan(threshold_value):
        axis.axhline(
            y=threshold_value,
            linestyle=":",
            linewidth=1.5,
            label=f"threshold = {threshold_value:.3g}",
        )

    maybe_symlog_x_axis(axis_values=plot_df["spike_n"], ax=axis)
    axis.set_xlabel("Spike level")
    axis.set_ylabel(metric.replace("_", " "))
    axis.set_title(f"{workflow}: {target_label} - raw replicate traces")
    axis.grid(True, alpha=0.3)
    axis.legend(loc="best", fontsize=7)
    figure.tight_layout()

    stem = sanitise_filename(
        value=f"trace__{workflow}__{target_label}__{metric}"
    )
    png_path = out_dir / f"{stem}.png"
    pdf_path = out_dir / f"{stem}.pdf"
    figure.savefig(fname=png_path, dpi=200)
    figure.savefig(fname=pdf_path)
    plt.close(fig=figure)

    return {
        "plot_type": "raw_replicate_traces",
        "workflow": workflow,
        "metric": metric,
        "target_label": target_label,
        "png_path": png_path.name,
        "pdf_path": pdf_path.name,
        "caption": (
            "One line per run and replicate. The dotted horizontal line marks "
            "the detection threshold."
        ),
    }


def generate_detection_heatmap(
    *,
    plot_df: pd.DataFrame,
    out_dir: Path,
    workflow: str,
    metric: str,
    target_label: str,
) -> Optional[dict[str, str]]:
    """
    Generate a replicate-by-spike detection heatmap.

    Parameters
    ----------
    plot_df : pd.DataFrame
        Subset for one workflow, metric, and target.
    out_dir : Path
        Plot output directory.
    workflow : str
        Workflow name.
    metric : str
        Metric name.
    target_label : str
        Target label.

    Returns
    -------
    Optional[dict[str, str]]
        Plot manifest entry.
    """
    if plot_df.empty:
        return None

    heatmap_df = plot_df.copy()
    heatmap_df["row_label"] = heatmap_df.apply(
        lambda row: make_trace_label(
            run_name=str(row["run_name"]),
            replicate=row["replicate"],
            is_shuffled=bool(row["is_shuffled_control"]),
        ),
        axis=1,
    )
    heatmap_df["detected_int"] = heatmap_df["detected"].astype(int)

    pivot = heatmap_df.pivot_table(
        index="row_label",
        columns="spike_n",
        values="detected_int",
        aggfunc="max",
        fill_value=0,
    )

    if pivot.empty:
        return None

    figure_height = max(3.5, 0.35 * pivot.shape[0])
    figure, axis = plt.subplots(figsize=(9, figure_height))
    image = axis.imshow(pivot.values, aspect="auto")
    axis.set_xticks(range(pivot.shape[1]))
    axis.set_xticklabels([str(column) for column in pivot.columns], rotation=45, ha="right")
    axis.set_yticks(range(pivot.shape[0]))
    axis.set_yticklabels(list(pivot.index))
    axis.set_xlabel("Spike level")
    axis.set_ylabel("Run and replicate")
    axis.set_title(f"{workflow}: {target_label} - detection heatmap ({metric})")
    figure.colorbar(image, ax=axis, fraction=0.03, pad=0.02)
    figure.tight_layout()

    stem = sanitise_filename(
        value=f"heatmap__{workflow}__{target_label}__{metric}"
    )
    png_path = out_dir / f"{stem}.png"
    pdf_path = out_dir / f"{stem}.pdf"
    figure.savefig(fname=png_path, dpi=200)
    figure.savefig(fname=pdf_path)
    plt.close(fig=figure)

    return {
        "plot_type": "replicate_detection_heatmap",
        "workflow": workflow,
        "metric": metric,
        "target_label": target_label,
        "png_path": png_path.name,
        "pdf_path": pdf_path.name,
        "caption": (
            "Rows are individual run and replicate combinations. Columns are "
            "spike levels. Cell values indicate whether that observation was "
            "called detected."
        ),
    }


def format_scalar(*, value: object) -> str:
    """
    Format a scalar for HTML display.

    Parameters
    ----------
    value : object
        Value to format.

    Returns
    -------
    str
        Formatted string.
    """
    if pd.isna(value):
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return f"{int(value):,}"
        if abs(value) >= 1000:
            return f"{value:,.3f}"
        return f"{value:.3f}"
    return html.escape(str(value))


def dataframe_to_html_table(
    *,
    dataframe: pd.DataFrame,
    max_rows: Optional[int] = None,
) -> str:
    """
    Render a dataframe as an HTML table.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Table to render.
    max_rows : Optional[int]
        Optional row cap for display.

    Returns
    -------
    str
        HTML table.
    """
    if dataframe is None or dataframe.empty:
        return '<p class="muted">No rows available.</p>'

    display_df = dataframe.copy()
    if max_rows is not None and display_df.shape[0] > max_rows:
        display_df = display_df.head(max_rows).copy()

    headers = "".join(f"<th>{html.escape(str(column))}</th>" for column in display_df.columns)
    body_rows: list[str] = []
    for _, row in display_df.iterrows():
        cells = "".join(f"<td>{format_scalar(value=row[column])}</td>" for column in display_df.columns)
        body_rows.append(f"<tr>{cells}</tr>")

    table = (
        '<div class="table-wrap"><table class="data-table">'
        f"<thead><tr>{headers}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody>"
        "</table></div>"
    )
    if max_rows is not None and dataframe.shape[0] > max_rows:
        table += (
            f'<p class="muted">Showing first {max_rows:,} of '
            f"{dataframe.shape[0]:,} rows in the HTML view.</p>"
        )
    return table


def build_summary_cards(*, entries: list[tuple[str, object]]) -> str:
    """
    Build HTML summary cards.

    Parameters
    ----------
    entries : list[tuple[str, object]]
        Label/value pairs.

    Returns
    -------
    str
        HTML snippet.
    """
    cards = []
    for label, value in entries:
        cards.append(
            "<div class=\"summary-card\">"
            f"<div class=\"summary-label\">{html.escape(str(label))}</div>"
            f"<div class=\"summary-value\">{html.escape(str(value))}</div>"
            "</div>"
        )
    return f'<div class="summary-grid">{"".join(cards)}</div>'


def build_plot_gallery(
    *,
    plot_manifest: pd.DataFrame,
    plot_type: str,
    relative_plot_dir: str,
) -> str:
    """
    Build an HTML plot gallery for one plot type.

    Parameters
    ----------
    plot_manifest : pd.DataFrame
        Plot manifest table.
    plot_type : str
        Plot type to filter.
    relative_plot_dir : str
        Relative plot directory path.

    Returns
    -------
    str
        HTML snippet.
    """
    subset = plot_manifest.loc[plot_manifest["plot_type"] == plot_type].copy()
    if subset.empty:
        return '<p class="muted">No plots available.</p>'

    cards: list[str] = []
    for _, row in subset.iterrows():
        png_href = f"{relative_plot_dir}/{row['png_path']}"
        caption = html.escape(str(row.get("caption", "")))
        title = html.escape(
            f"{row['workflow']} | {row['target_label']} | {row['metric']}"
        )
        cards.append(
            "<figure class=\"plot-card\">"
            f"<a href=\"{png_href}\" target=\"_blank\">"
            f"<img src=\"{png_href}\" alt=\"{title}\"></a>"
            f"<figcaption><strong>{title}</strong><br>{caption}</figcaption>"
            "</figure>"
        )
    return f'<div class="plot-grid">{"".join(cards)}</div>'


def format_excel_workbook(
    *,
    workbook_path: Path,
    percentage_columns: set[str],
    integer_columns: set[str],
) -> None:
    """
    Apply workbook styling similar to the existing summary output.

    Parameters
    ----------
    workbook_path : Path
        Workbook path.
    percentage_columns : set[str]
        Columns shown with three decimals.
    integer_columns : set[str]
        Columns shown as integers.
    """
    workbook = load_workbook(filename=workbook_path)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(vertical="center", wrap_text=True)

    for worksheet in workbook.worksheets:
        if worksheet.max_row == 0 or worksheet.max_column == 0:
            continue

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        headers = [cell.value for cell in worksheet[1]]
        for idx, header in enumerate(headers, start=1):
            if header is None:
                continue
            header_text = str(header)
            column_letter = get_column_letter(idx)
            max_len = len(header_text)
            for row in worksheet.iter_rows(
                min_row=2,
                min_col=idx,
                max_col=idx,
                values_only=False,
            ):
                cell = row[0]
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
                    if header_text in percentage_columns and isinstance(cell.value, (int, float)):
                        cell.number_format = "0.000"
                    elif header_text in integer_columns and isinstance(cell.value, (int, float)):
                        cell.number_format = "0"
            worksheet.column_dimensions[column_letter].width = min(max_len + 2, 42)

    workbook.save(filename=workbook_path)


def main() -> None:
    """
    Run the replicate-resolved reporting workflow.
    """
    args = parse_args()

    summary_dir = Path(args.summary_dir).expanduser().resolve()
    if not summary_dir.exists():
        raise FileNotFoundError(f"Summary directory not found: {summary_dir}")

    out_dir = (
        Path(args.out_dir).expanduser().resolve()
        if args.out_dir
        else summary_dir / "replicate_resolved_report"
    )
    out_dir.mkdir(parents=True, exist_ok=True)
    plots_dir = out_dir / "plots"
    plots_dir.mkdir(parents=True, exist_ok=True)

    combined_long_path = summary_dir / "combined_long.tsv"
    if not combined_long_path.exists():
        raise FileNotFoundError(f"Required file not found: {combined_long_path}")

    combined_long = pd.read_csv(filepath_or_buffer=combined_long_path, sep="\t")
    combined_long = normalise_combined_long(combined_long=combined_long)

    if args.spike_axis == "per_genome":
        positive_column = "is_positive_per_genome"
        negative_column = "is_negative_per_genome"
        spike_column = "spike_n_per_genome"
    else:
        positive_column = "is_positive_total"
        negative_column = "is_negative_total"
        spike_column = "total_spike_n"

    detection_df, method_threshold_df = compute_detection_calls(
        dataframe=combined_long,
        group_columns=["workflow", "metric", "target_label"],
        value_column="value",
        positive_column=positive_column,
        negative_column=negative_column,
        spike_column=spike_column,
        threshold_mode=args.threshold_mode,
        min_detect_value=args.min_detect_value,
        sd_multiplier=args.sd_multiplier,
    )

    method_performance_df = summarise_performance(
        detection_df=detection_df,
        group_columns=["workflow", "metric", "target_label"],
    )
    run_performance_df = summarise_performance(
        detection_df=detection_df,
        group_columns=["workflow", "run_name", "metric", "target_label"],
    )
    replicate_performance_df = summarise_performance(
        detection_df=detection_df,
        group_columns=["workflow", "run_name", "replicate", "metric", "target_label"],
    )

    method_by_spike_df = summarise_detection_by_spike(
        detection_df=detection_df,
        group_columns=["workflow", "metric", "target_label"],
    )
    run_by_spike_df = summarise_detection_by_spike(
        detection_df=detection_df,
        group_columns=["workflow", "run_name", "metric", "target_label"],
    )
    signal_by_spike_df = summarise_signal_by_spike(
        dataframe=detection_df,
        value_column="value",
        group_columns=["workflow", "run_name", "metric", "target_label"],
        include_negatives=True,
    )
    replicate_first_detection_df = summarise_replicate_first_detection(
        detection_df=detection_df,
    )

    method_threshold_summary_df = summarise_thresholds_with_first_detection(
        threshold_df=method_threshold_df,
        by_spike_df=method_by_spike_df,
        group_columns=["workflow", "metric", "target_label"],
    )
    run_threshold_summary_df = summarise_thresholds_with_first_detection(
        threshold_df=run_performance_df[[
            "workflow",
            "run_name",
            "metric",
            "target_label",
            "threshold_mode",
            "threshold_value",
            "negative_mean",
            "negative_max",
            "negative_sd",
            "n_negative_for_threshold",
        ]].drop_duplicates(),
        by_spike_df=run_by_spike_df,
        group_columns=["workflow", "run_name", "metric", "target_label"],
    )

    plot_records: list[dict[str, str]] = []
    grouped = detection_df.groupby(["workflow", "metric", "target_label"], dropna=False)
    for keys, panel_df in grouped:
        workflow, metric, target_label = keys
        panel_threshold = float(panel_df["threshold_value"].iloc[0])
        raw_plot = generate_raw_signal_plot(
            plot_df=panel_df,
            threshold_value=panel_threshold,
            out_dir=plots_dir,
            workflow=str(workflow),
            metric=str(metric),
            target_label=str(target_label),
        )
        if raw_plot is not None:
            plot_records.append(raw_plot)
        heatmap_plot = generate_detection_heatmap(
            plot_df=panel_df,
            out_dir=plots_dir,
            workflow=str(workflow),
            metric=str(metric),
            target_label=str(target_label),
        )
        if heatmap_plot is not None:
            plot_records.append(heatmap_plot)

    plot_manifest_df = pd.DataFrame(plot_records)

    reported_taxa_path = summary_dir / "reported_taxa_long.tsv"
    reported_taxa_detection_df = pd.DataFrame()
    reported_taxa_threshold_df = pd.DataFrame()
    reported_taxa_found_df = pd.DataFrame()
    if reported_taxa_path.exists() and reported_taxa_path.stat().st_size > 0:
        reported_taxa_long = pd.read_csv(filepath_or_buffer=reported_taxa_path, sep="\t")
        reported_taxa_long = normalise_reported_taxa_long(dataframe=reported_taxa_long)
        reported_taxa_long = choose_reported_taxa_metric(dataframe=reported_taxa_long)
        reported_taxa_detection_df, reported_taxa_threshold_df = compute_detection_calls(
            dataframe=reported_taxa_long,
            group_columns=["workflow", "classifier", "taxon_name", "metric_name"],
            value_column="metric_value",
            positive_column="is_positive",
            negative_column="is_negative",
            spike_column="spike_n",
            threshold_mode=args.threshold_mode,
            min_detect_value=args.min_detect_value,
            sd_multiplier=args.sd_multiplier,
        )
        reported_taxa_found_df = summarise_reported_taxa_found(
            detection_df=reported_taxa_detection_df,
        )

    output_tables: dict[str, pd.DataFrame] = {
        "detection_calls": detection_df,
        "method_thresholds": method_threshold_summary_df,
        "method_performance": method_performance_df,
        "run_thresholds": run_threshold_summary_df,
        "run_performance": run_performance_df,
        "replicate_performance": replicate_performance_df,
        "replicate_first_detection": replicate_first_detection_df,
        "method_by_spike": method_by_spike_df,
        "run_by_spike": run_by_spike_df,
        "signal_by_spike": signal_by_spike_df,
        "plot_manifest": plot_manifest_df,
    }

    if not reported_taxa_detection_df.empty:
        output_tables["reported_taxa_detection_calls"] = reported_taxa_detection_df
    if not reported_taxa_threshold_df.empty:
        output_tables["reported_taxa_thresholds"] = reported_taxa_threshold_df
    if not reported_taxa_found_df.empty:
        output_tables["reported_taxa_found"] = reported_taxa_found_df

    for stem, dataframe in output_tables.items():
        dataframe.to_csv(path_or_buf=out_dir / f"{stem}.tsv", sep="\t", index=False)

    workbook_path = out_dir / "replicate_resolved_report.xlsx"
    with pd.ExcelWriter(path=workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "metric": [
                    "n_observations",
                    "n_methods",
                    "n_run_metric_panels",
                    "n_run_replicate_panels",
                    "n_plots",
                    "threshold_mode",
                    "spike_axis",
                ],
                "value": [
                    int(detection_df.shape[0]),
                    int(method_threshold_summary_df.shape[0]),
                    int(run_performance_df.shape[0]),
                    int(replicate_performance_df.shape[0]),
                    int(plot_manifest_df.shape[0]),
                    args.threshold_mode,
                    args.spike_axis,
                ],
            }
        ).to_excel(writer, sheet_name="executive_summary", index=False)
        method_threshold_summary_df.to_excel(writer, sheet_name="method_thresholds", index=False)
        method_performance_df.to_excel(writer, sheet_name="method_performance", index=False)
        run_threshold_summary_df.to_excel(writer, sheet_name="run_thresholds", index=False)
        run_performance_df.to_excel(writer, sheet_name="run_performance", index=False)
        replicate_performance_df.to_excel(writer, sheet_name="replicate_perf", index=False)
        replicate_first_detection_df.to_excel(writer, sheet_name="replicate_first", index=False)
        method_by_spike_df.to_excel(writer, sheet_name="method_by_spike", index=False)
        run_by_spike_df.to_excel(writer, sheet_name="run_by_spike", index=False)
        signal_by_spike_df.to_excel(writer, sheet_name="signal_by_spike", index=False)
        detection_df.to_excel(writer, sheet_name="detection_calls", index=False)
        plot_manifest_df.to_excel(writer, sheet_name="plot_manifest", index=False)
        if not reported_taxa_threshold_df.empty:
            reported_taxa_threshold_df.to_excel(writer, sheet_name="reported_thresholds", index=False)
        if not reported_taxa_found_df.empty:
            reported_taxa_found_df.to_excel(writer, sheet_name="reported_taxa_found", index=False)
        if not reported_taxa_detection_df.empty:
            reported_taxa_detection_df.to_excel(writer, sheet_name="reported_taxa_calls", index=False)

    percentage_columns = {
        "sensitivity",
        "specificity",
        "precision",
        "negative_predictive_value",
        "accuracy",
        "balanced_accuracy",
        "false_positive_rate",
        "false_negative_rate",
        "f1_score",
        "youden_j",
        "detection_rate",
        "detection_rate_ci_low",
        "detection_rate_ci_high",
        "mean",
        "median",
        "std",
        "min",
        "max",
        "threshold_value",
        "negative_mean",
        "negative_max",
        "negative_sd",
        "value",
        "metric_value",
        "max_signal",
    }
    integer_columns = {
        "replicate",
        "tp",
        "tn",
        "fp",
        "fn",
        "n_positive",
        "n_negative",
        "n_negative_for_threshold",
        "n_observations",
        "n_detected",
        "n_positive_spikes",
        "n_detected_positive_spikes",
        "spike_n",
        "first_detected_spike",
        "first_detected_spike_any",
        "first_detected_spike_50pct",
        "first_detected_spike_95pct",
        "first_detected_spike_all",
        "lod50_spike_n",
        "lod95_spike_n",
        "lod100_spike_n",
        "total_spike_n",
        "spike_n_per_genome",
        "n_taxa_found",
    }
    format_excel_workbook(
        workbook_path=workbook_path,
        percentage_columns=percentage_columns,
        integer_columns=integer_columns,
    )

    cards_html = build_summary_cards(
        entries=[
            ("Observations", f"{detection_df.shape[0]:,}"),
            ("Methods", f"{method_threshold_summary_df.shape[0]:,}"),
            ("Run-level panels", f"{run_performance_df.shape[0]:,}"),
            ("Replicate-level panels", f"{replicate_performance_df.shape[0]:,}"),
            ("Plots", f"{plot_manifest_df.shape[0]:,}"),
            ("Threshold rule", args.threshold_mode),
            ("Spike axis", args.spike_axis),
        ]
    )

    html_page = f"""<!DOCTYPE html>
<html lang=\"en\">
<head>
  <meta charset=\"utf-8\">
  <title>{html.escape(args.title)}</title>
  {HTML_STYLE}
</head>
<body>
  <div class=\"container\">
    <h1>{html.escape(args.title)}</h1>
    <p class=\"lede\">
      Replicate-resolved benchmarking report for the ONT spike-in framework.
      This view keeps every run and replicate separate so that partial
      detection at a given spike level can be inspected directly rather than
      being visually collapsed into a single median line.
    </p>

    <div class=\"note-box\">
      <strong>How to use this report:</strong>
      Start with the threshold and run-level tables to see the raw metric cut
      offs and the first spike at which detections begin. Then use the raw
      trace plots and heatmaps to inspect whether all replicates behaved
      consistently for each workflow, metric, and target.
    </div>

    <section class=\"report-section\">
      <h2>Executive summary</h2>
      {cards_html}
    </section>

    <section class=\"report-section\">
      <h2>Method thresholds and first detection summary</h2>
      <p class=\"muted\">
        This table separates the raw signal threshold from the spike-level
        summary. Threshold values are derived from negative observations using
        the selected threshold rule. The first-detection columns show the first
        spike level at which any replicate, at least half of replicates,
        at least 95 percent of replicates, or all replicates were called
        detected.
      </p>
      {dataframe_to_html_table(dataframe=method_threshold_summary_df, max_rows=200)}
    </section>

    <section class=\"report-section\">
      <h2>Run-level thresholds and first detection summary</h2>
      {dataframe_to_html_table(dataframe=run_threshold_summary_df, max_rows=300)}
    </section>

    <section class=\"report-section\">
      <h2>Run-level performance statistics</h2>
      {dataframe_to_html_table(dataframe=run_performance_df, max_rows=300)}
    </section>

    <section class=\"report-section\">
      <h2>Replicate-level first detection summary</h2>
      {dataframe_to_html_table(dataframe=replicate_first_detection_df, max_rows=400)}
    </section>

    <section class=\"report-section\">
      <h2>Detection rate by spike level</h2>
      {dataframe_to_html_table(dataframe=run_by_spike_df, max_rows=400)}
    </section>

    <section class=\"report-section\">
      <h2>Raw signal summary by spike level</h2>
      {dataframe_to_html_table(dataframe=signal_by_spike_df, max_rows=400)}
    </section>

    <section class=\"report-section\">
      <h2>Replicate raw-signal trace plots</h2>
      {build_plot_gallery(plot_manifest=plot_manifest_df, plot_type='raw_replicate_traces', relative_plot_dir='plots')}
    </section>

    <section class=\"report-section\">
      <h2>Replicate detection heatmaps</h2>
      {build_plot_gallery(plot_manifest=plot_manifest_df, plot_type='replicate_detection_heatmap', relative_plot_dir='plots')}
    </section>

    <section class=\"report-section\">
      <h2>Reported taxa above threshold</h2>
      {dataframe_to_html_table(dataframe=reported_taxa_found_df, max_rows=300)}
    </section>

    <details>
      <summary>Show full observation-level detection call table</summary>
      {dataframe_to_html_table(dataframe=detection_df, max_rows=500)}
    </details>
  </div>
</body>
</html>
"""

    html_path = out_dir / "replicate_resolved_report.html"
    html_path.write_text(data=html_page, encoding="utf-8")


if __name__ == "__main__":
    main()
