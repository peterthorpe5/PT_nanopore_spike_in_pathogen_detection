#!/usr/bin/env python3
"""
Build method-performance summary tables from the spike-in combined long table.

This script converts the long-format benchmarking table produced by the
spike-in summariser into compact method-level performance tables suitable
for inclusion in the HTML report.

It treats each workflow / metric / target-label combination as a separate
benchmarking method and calculates confusion-matrix-style performance
statistics using a configurable detection threshold.

Compared with earlier versions, this script also:
- writes a formatted Excel workbook
- produces a more readable HTML page and HTML fragment
- includes a compact headline table and a full-detail table
- includes definitions of all reported metrics
- includes notes on how to interpret the plots in the main HTML report
- ranks methods in a more useful order for detection benchmarking
- adds a panel-level species summary table
- adds a plain-language interpretation column

Default interpretation
----------------------
- Positive observations:
  non-shuffled rows with spike_n > 0
- Negative observations:
  shuffled-control rows or rows with spike_n == 0

Default detection rule
----------------------
Detection is called when the observed metric value is greater than or equal
to the method-specific maximum observed negative value, with a floor set by
--min_detect_value. This is intentionally conservative.
"""

from __future__ import annotations

import argparse
import html
import math
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


METRIC_DEFINITIONS = [
    {
        "term": "TP",
        "full_name": "True positives",
        "definition": "Positive samples that were correctly called detected.",
        "formula": "Count of positive observations with detected = True",
    },
    {
        "term": "FP",
        "full_name": "False positives",
        "definition": "Negative samples that were incorrectly called detected.",
        "formula": "Count of negative observations with detected = True",
    },
    {
        "term": "TN",
        "full_name": "True negatives",
        "definition": "Negative samples that were correctly called not detected.",
        "formula": "Count of negative observations with detected = False",
    },
    {
        "term": "FN",
        "full_name": "False negatives",
        "definition": "Positive samples that were incorrectly called not detected.",
        "formula": "Count of positive observations with detected = False",
    },
    {
        "term": "sensitivity",
        "full_name": "Sensitivity",
        "definition": (
            "Proportion of true positives recovered among all positive samples."
        ),
        "formula": "TP / (TP + FN)",
    },
    {
        "term": "recall",
        "full_name": "Recall",
        "definition": "Same quantity as sensitivity in this framework.",
        "formula": "TP / (TP + FN)",
    },
    {
        "term": "true_positive_rate",
        "full_name": "True positive rate",
        "definition": "Same quantity as sensitivity and recall in this framework.",
        "formula": "TP / (TP + FN)",
    },
    {
        "term": "specificity",
        "full_name": "Specificity",
        "definition": "Proportion of true negatives correctly rejected.",
        "formula": "TN / (TN + FP)",
    },
    {
        "term": "false_positive_rate",
        "full_name": "False positive rate",
        "definition": (
            "Proportion of negative samples incorrectly called positive."
        ),
        "formula": "FP / (FP + TN)",
    },
    {
        "term": "false_negative_rate",
        "full_name": "False negative rate",
        "definition": "Proportion of positive samples missed by the method.",
        "formula": "FN / (FN + TP)",
    },
    {
        "term": "precision",
        "full_name": "Precision",
        "definition": "Among all positive calls, the proportion that were correct.",
        "formula": "TP / (TP + FP)",
    },
    {
        "term": "negative_predictive_value",
        "full_name": "Negative predictive value",
        "definition": "Among all negative calls, the proportion that were correct.",
        "formula": "TN / (TN + FN)",
    },
    {
        "term": "accuracy",
        "full_name": "Accuracy",
        "definition": "Overall proportion of correct calls.",
        "formula": "(TP + TN) / (TP + TN + FP + FN)",
    },
    {
        "term": "balanced_accuracy",
        "full_name": "Balanced accuracy",
        "definition": (
            "Average of sensitivity and specificity. Useful when positive and "
            "negative classes are imbalanced."
        ),
        "formula": "(Sensitivity + Specificity) / 2",
    },
    {
        "term": "f1_score",
        "full_name": "F1 score",
        "definition": "Harmonic mean of precision and recall.",
        "formula": "2TP / (2TP + FP + FN)",
    },
    {
        "term": "youden_j",
        "full_name": "Youden's J",
        "definition": (
            "Single summary index combining sensitivity and specificity. "
            "Higher values indicate better separation from negatives."
        ),
        "formula": "Sensitivity + Specificity - 1",
    },
    {
        "term": "lod50_spike_n",
        "full_name": "LOD50 by spike level",
        "definition": (
            "Lowest spike_n at which the detection rate is at least 50 percent."
        ),
        "formula": (
            "First spike_n where mean(detected) across positive observations "
            "is >= 0.50"
        ),
    },
    {
        "term": "lod95_spike_n",
        "full_name": "LOD95 by spike level",
        "definition": (
            "Lowest spike_n at which the detection rate is at least 95 percent."
        ),
        "formula": (
            "First spike_n where mean(detected) across positive observations "
            "is >= 0.95"
        ),
    },
]

REPORT_PLOT_EXPLANATIONS = [
    {
        "plot_name": "Detection overview plots",
        "description": (
            "These plots show how each method's raw signal changes as more "
            "simulated pathogen reads are added. The x-axis is total spiked "
            "reads and the y-axis is the raw method-specific signal, such as "
            "Kraken target reads, minimap target alignments, MetaMaps target "
            "reads, or Kraken target contigs."
        ),
    },
    {
        "plot_name": "Separate panels by workflow, metric, and target",
        "description": (
            "Each panel corresponds to one workflow, one metric, and one target "
            "label. This makes it easier to compare how a given target behaves "
            "within a method without mixing together signals from different "
            "tools or species."
        ),
    },
    {
        "plot_name": "Lines within a panel",
        "description": (
            "Lines represent different run families within the same workflow. "
            "The plotted value is the median signal at each spike level across "
            "replicates within that run."
        ),
    },
    {
        "plot_name": "Shuffled controls",
        "description": (
            "Any line labelled as shuffled is a randomised negative control. "
            "Ideally, these lines should remain close to baseline. Upward "
            "movement in shuffled controls suggests background noise or "
            "false-positive behaviour."
        ),
    },
    {
        "plot_name": "How to interpret good performance",
        "description": (
            "A strong method typically shows low or stable signal in negatives "
            "and shuffled controls, increasing signal as spike level rises, "
            "and earlier separation from baseline at low abundance."
        ),
    },
    {
        "plot_name": "What the performance table adds",
        "description": (
            "The performance table turns those trends into summary statistics "
            "using a configurable detection threshold. It gives an at-a-glance "
            "comparison of sensitivity, specificity, precision, false-positive "
            "rate, and limit-of-detection style summaries."
        ),
    },
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build method-level performance summary tables from a combined "
            "long-format spike-in benchmarking table."
        )
    )
    parser.add_argument("--combined_long_tsv", required=True)
    parser.add_argument("--out_dir", required=True)
    parser.add_argument(
        "--threshold_mode",
        choices=["fixed", "baseline_mean", "baseline_max", "baseline_mean_plus_sd"],
        default="baseline_max",
    )
    parser.add_argument("--min_detect_value", type=float, default=1.0)
    parser.add_argument("--sd_multiplier", type=float, default=2.0)
    return parser.parse_args()


def get_first_existing_column(dataframe: pd.DataFrame, candidates: list[str]) -> Optional[str]:
    for candidate in candidates:
        if candidate in dataframe.columns:
            return candidate
    return None


def normalise_combined_long(combined_long: pd.DataFrame) -> pd.DataFrame:
    dataframe = combined_long.copy()
    rename_map: dict[str, str] = {}

    workflow_column = get_first_existing_column(dataframe, ["workflow", "workflow_type"])
    metric_column = get_first_existing_column(dataframe, ["metric", "metric_name"])
    value_column = get_first_existing_column(dataframe, ["value", "metric_value"])
    spike_column = get_first_existing_column(dataframe, ["spike_n", "spike_n_per_genome"])
    shuffled_column = get_first_existing_column(
        dataframe,
        ["is_shuffled_control_recomputed", "is_shuffled_control"],
    )
    run_name_column = get_first_existing_column(dataframe, ["run_name"])

    if workflow_column is None or metric_column is None or value_column is None or spike_column is None:
        raise ValueError("Missing one or more required columns in combined long table.")

    if workflow_column != "workflow":
        rename_map[workflow_column] = "workflow"
    if metric_column != "metric":
        rename_map[metric_column] = "metric"
    if value_column != "value":
        rename_map[value_column] = "value"
    if spike_column != "spike_n":
        rename_map[spike_column] = "spike_n"
    if shuffled_column and shuffled_column != "is_shuffled_control":
        rename_map[shuffled_column] = "is_shuffled_control"
    if run_name_column and run_name_column != "run_name":
        rename_map[run_name_column] = "run_name"

    if rename_map:
        dataframe = dataframe.rename(columns=rename_map)

    if "target_label" not in dataframe.columns:
        dataframe["target_label"] = "single_target"
    if "is_shuffled_control" not in dataframe.columns:
        dataframe["is_shuffled_control"] = False
    if "replicate" not in dataframe.columns:
        dataframe["replicate"] = pd.NA
    if "run_name" not in dataframe.columns:
        dataframe["run_name"] = dataframe["workflow"].astype(str)

    dataframe["workflow"] = dataframe["workflow"].fillna("unknown").astype(str)
    dataframe["metric"] = dataframe["metric"].fillna("unknown_metric").astype(str)
    dataframe["target_label"] = (
        dataframe["target_label"].fillna("single_target").replace("", "single_target").astype(str)
    )
    dataframe["run_name"] = dataframe["run_name"].fillna("unknown_run").astype(str)
    dataframe["value"] = pd.to_numeric(dataframe["value"], errors="coerce")
    dataframe["spike_n"] = pd.to_numeric(dataframe["spike_n"], errors="coerce")
    dataframe["is_shuffled_control"] = (
        dataframe["is_shuffled_control"].fillna(False).astype(str).str.lower().isin(["true", "1", "yes"])
    )
    dataframe = dataframe.dropna(subset=["value", "spike_n"]).copy()
    dataframe["is_positive"] = (~dataframe["is_shuffled_control"]) & (dataframe["spike_n"] > 0)
    dataframe["is_negative"] = dataframe["is_shuffled_control"] | (dataframe["spike_n"] == 0)
    return dataframe


def safe_divide(numerator: float, denominator: float) -> float:
    if denominator == 0:
        return math.nan
    return numerator / denominator


def choose_threshold(
    negative_values: pd.Series,
    threshold_mode: str,
    min_detect_value: float,
    sd_multiplier: float,
) -> float:
    negative_values = pd.to_numeric(negative_values, errors="coerce").dropna()
    if negative_values.empty:
        return float(min_detect_value)
    if threshold_mode == "fixed":
        return float(min_detect_value)
    if threshold_mode == "baseline_mean":
        return float(max(min_detect_value, negative_values.mean()))
    if threshold_mode == "baseline_max":
        return float(max(min_detect_value, negative_values.max()))
    if threshold_mode == "baseline_mean_plus_sd":
        threshold = negative_values.mean() + (sd_multiplier * negative_values.std(ddof=0))
        return float(max(min_detect_value, threshold))
    return float(min_detect_value)


def compute_detection_calls(
    dataframe: pd.DataFrame,
    threshold_mode: str,
    min_detect_value: float,
    sd_multiplier: float,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    detection_rows: list[pd.DataFrame] = []
    threshold_rows: list[dict[str, object]] = []

    grouped = dataframe.groupby(["workflow", "metric", "target_label"], dropna=False)
    for (workflow, metric, target_label), group in grouped:
        negatives = group.loc[group["is_negative"], "value"]
        threshold_value = choose_threshold(
            negative_values=negatives,
            threshold_mode=threshold_mode,
            min_detect_value=min_detect_value,
            sd_multiplier=sd_multiplier,
        )
        group = group.copy()
        group["threshold_value"] = threshold_value
        group["detected"] = group["value"] >= threshold_value
        detection_rows.append(group)
        threshold_rows.append(
            {
                "workflow": workflow,
                "metric": metric,
                "target_label": target_label,
                "threshold_mode": threshold_mode,
                "threshold_value": threshold_value,
                "negative_mean": float(negatives.mean()) if not negatives.dropna().empty else math.nan,
                "negative_max": float(negatives.max()) if not negatives.dropna().empty else math.nan,
                "negative_sd": float(negatives.std(ddof=0)) if not negatives.dropna().empty else math.nan,
                "n_negative_for_threshold": int(negatives.dropna().shape[0]),
            }
        )

    detection_df = pd.concat(detection_rows, ignore_index=True) if detection_rows else pd.DataFrame()
    threshold_df = pd.DataFrame(threshold_rows)
    return detection_df, threshold_df


def first_spike_meeting_rate(dataframe: pd.DataFrame, min_rate: float) -> float:
    if dataframe.empty:
        return math.nan
    grouped = (
        dataframe.groupby("spike_n", dropna=False)["detected"].mean().reset_index().sort_values(by="spike_n")
    )
    detected = grouped.loc[grouped["detected"] >= min_rate, "spike_n"]
    if detected.empty:
        return math.nan
    return float(detected.iloc[0])


def build_interpretation(row: pd.Series) -> str:
    sens = row.get("sensitivity", math.nan)
    spec = row.get("specificity", math.nan)
    fpr = row.get("false_positive_rate", math.nan)
    lod50 = row.get("lod50_spike_n", math.nan)

    if pd.isna(sens) or pd.isna(spec):
        return "Insufficient data for stable interpretation"
    if sens >= 0.90 and spec >= 0.90:
        if not pd.isna(lod50) and lod50 <= 10:
            return "Strong overall balance with early detection"
        return "Strong overall balance"
    if sens >= 0.90 and spec < 0.50:
        return "High sensitivity but poor negative discrimination"
    if sens < 0.50 and spec >= 0.90:
        return "Very conservative and misses many positives"
    if sens < 0.50 and spec < 0.50:
        return "Weak on both positives and negatives"
    if not pd.isna(fpr) and fpr == 0:
        return "No observed false positives, but sensitivity is limited"
    return "Intermediate performance with trade-offs"


def summarise_method_performance(
    detection_df: pd.DataFrame,
    threshold_df: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    grouped = detection_df.groupby(["workflow", "metric", "target_label"], dropna=False)

    for (workflow, metric, target_label), group in grouped:
        positives = group.loc[group["is_positive"]].copy()
        negatives = group.loc[group["is_negative"]].copy()

        tp = int((positives["detected"] == True).sum())
        fn = int((positives["detected"] == False).sum())
        fp = int((negatives["detected"] == True).sum())
        tn = int((negatives["detected"] == False).sum())

        sensitivity = safe_divide(tp, tp + fn)
        recall = sensitivity
        true_positive_rate = sensitivity
        specificity = safe_divide(tn, tn + fp)
        false_positive_rate = safe_divide(fp, fp + tn)
        false_negative_rate = safe_divide(fn, fn + tp)
        precision = safe_divide(tp, tp + fp)
        negative_predictive_value = safe_divide(tn, tn + fn)
        accuracy = safe_divide(tp + tn, tp + tn + fp + fn)
        balanced_accuracy = (
            (sensitivity + specificity) / 2
            if not math.isnan(sensitivity) and not math.isnan(specificity)
            else math.nan
        )
        f1_score = safe_divide(2 * tp, (2 * tp) + fp + fn)
        youden_j = (
            sensitivity + specificity - 1
            if not math.isnan(sensitivity) and not math.isnan(specificity)
            else math.nan
        )

        lod_50 = first_spike_meeting_rate(positives, 0.50)
        lod_95 = first_spike_meeting_rate(positives, 0.95)

        threshold_row = threshold_df.loc[
            (threshold_df["workflow"] == workflow)
            & (threshold_df["metric"] == metric)
            & (threshold_df["target_label"] == target_label)
        ]
        threshold_record = threshold_row.iloc[0].to_dict() if not threshold_row.empty else {}

        rows.append(
            {
                "workflow": workflow,
                "metric": metric,
                "target_label": target_label,
                "n_positive": int(positives.shape[0]),
                "n_negative": int(negatives.shape[0]),
                "tp": tp,
                "fn": fn,
                "fp": fp,
                "tn": tn,
                "threshold_mode": threshold_record.get("threshold_mode", pd.NA),
                "threshold_value": threshold_record.get("threshold_value", math.nan),
                "negative_mean": threshold_record.get("negative_mean", math.nan),
                "negative_max": threshold_record.get("negative_max", math.nan),
                "negative_sd": threshold_record.get("negative_sd", math.nan),
                "sensitivity": sensitivity,
                "recall": recall,
                "true_positive_rate": true_positive_rate,
                "specificity": specificity,
                "false_positive_rate": false_positive_rate,
                "false_negative_rate": false_negative_rate,
                "precision": precision,
                "negative_predictive_value": negative_predictive_value,
                "accuracy": accuracy,
                "balanced_accuracy": balanced_accuracy,
                "f1_score": f1_score,
                "youden_j": youden_j,
                "lod50_spike_n": lod_50,
                "lod95_spike_n": lod_95,
            }
        )

    out_df = pd.DataFrame(rows)
    if not out_df.empty:
        out_df["interpretation"] = out_df.apply(build_interpretation, axis=1)
        out_df = out_df.sort_values(by=["workflow", "metric", "target_label"]).reset_index(drop=True)
    return out_df


def summarise_detection_by_spike(detection_df: pd.DataFrame) -> pd.DataFrame:
    positive_df = detection_df.loc[detection_df["is_positive"]].copy()
    if positive_df.empty:
        return pd.DataFrame(
            columns=[
                "workflow",
                "metric",
                "target_label",
                "spike_n",
                "n_observations",
                "n_detected",
                "detection_rate",
            ]
        )
    grouped = (
        positive_df.groupby(["workflow", "metric", "target_label", "spike_n"], dropna=False)
        .agg(
            n_observations=("detected", "size"),
            n_detected=("detected", "sum"),
            detection_rate=("detected", "mean"),
        )
        .reset_index()
        .sort_values(by=["workflow", "metric", "target_label", "spike_n"])
        .reset_index(drop=True)
    )
    return grouped


def rank_methods(method_df: pd.DataFrame) -> pd.DataFrame:
    if method_df.empty:
        return method_df.copy()
    ranked = method_df.copy()
    ranked["lod50_rank"] = ranked["lod50_spike_n"].fillna(10**12)
    ranked["lod95_rank"] = ranked["lod95_spike_n"].fillna(10**12)
    ranked = ranked.sort_values(
        by=[
            "balanced_accuracy",
            "specificity",
            "false_positive_rate",
            "sensitivity",
            "lod50_rank",
            "lod95_rank",
            "precision",
            "f1_score",
        ],
        ascending=[False, False, True, False, True, True, False, False],
    ).reset_index(drop=True)
    ranked["rank"] = range(1, len(ranked) + 1)
    return ranked.drop(columns=["lod50_rank", "lod95_rank"])


def build_compact_table(method_df: pd.DataFrame) -> pd.DataFrame:
    if method_df.empty:
        return pd.DataFrame()
    ranked = rank_methods(method_df)
    compact = ranked[
        [
            "rank",
            "workflow",
            "metric",
            "target_label",
            "threshold_value",
            "sensitivity",
            "specificity",
            "precision",
            "f1_score",
            "false_positive_rate",
            "false_negative_rate",
            "lod50_spike_n",
            "lod95_spike_n",
            "tp",
            "fn",
            "fp",
            "tn",
            "interpretation",
        ]
    ].copy()
    return compact.rename(
        columns={
            "rank": "Rank",
            "workflow": "Workflow",
            "metric": "Metric",
            "target_label": "Target",
            "threshold_value": "Threshold",
            "sensitivity": "Sensitivity",
            "specificity": "Specificity",
            "precision": "Precision",
            "f1_score": "F1",
            "false_positive_rate": "False positive rate",
            "false_negative_rate": "False negative rate",
            "lod50_spike_n": "LOD50",
            "lod95_spike_n": "LOD95",
            "tp": "TP",
            "fn": "FN",
            "fp": "FP",
            "tn": "TN",
            "interpretation": "Interpretation",
        }
    )


def build_panel_species_summary(
    detection_df: pd.DataFrame,
    threshold_df: pd.DataFrame,
) -> pd.DataFrame:
    positive_df = detection_df.loc[detection_df["is_positive"]].copy()
    if positive_df.empty:
        return pd.DataFrame()

    rows: list[dict[str, object]] = []
    grouped = positive_df.groupby(["workflow", "metric", "run_name"], dropna=False)

    for (workflow, metric, run_name), panel_df in grouped:
        expected_species = sorted(pd.unique(panel_df["target_label"].astype(str)))
        found_species: list[str] = []
        threshold_values: list[float] = []
        per_species_summaries: list[str] = []

        for species in expected_species:
            species_df = panel_df.loc[panel_df["target_label"] == species].copy()
            threshold_row = threshold_df.loc[
                (threshold_df["workflow"] == workflow)
                & (threshold_df["metric"] == metric)
                & (threshold_df["target_label"] == species)
            ]
            threshold_value = float(threshold_row.iloc[0]["threshold_value"]) if not threshold_row.empty else math.nan
            if not math.isnan(threshold_value):
                threshold_values.append(threshold_value)

            tp = int((species_df["detected"] == True).sum())
            fn = int((species_df["detected"] == False).sum())
            if tp > 0:
                found_species.append(species)

            sensitivity = safe_divide(tp, tp + fn)
            per_species_summaries.append(
                f"{species}: TP={tp}, FN={fn}, sensitivity={format_value(sensitivity)}"
            )

        n_expected = len(expected_species)
        n_found = len(found_species)
        missing_species = [x for x in expected_species if x not in found_species]
        panel_sensitivity = safe_divide(n_found, n_expected)

        negative_panel = detection_df.loc[
            (detection_df["workflow"] == workflow)
            & (detection_df["metric"] == metric)
            & (detection_df["run_name"] == run_name)
            & (detection_df["is_negative"])
        ].copy()
        fp = int((negative_panel["detected"] == True).sum())
        tn = int((negative_panel["detected"] == False).sum())
        specificity = safe_divide(tn, tn + fp)
        false_positive_rate = safe_divide(fp, fp + tn)

        lod50_values: list[float] = []
        lod95_values: list[float] = []
        for species in expected_species:
            species_df = panel_df.loc[panel_df["target_label"] == species].copy()
            lod50_val = first_spike_meeting_rate(species_df, 0.50)
            lod95_val = first_spike_meeting_rate(species_df, 0.95)
            if not math.isnan(lod50_val):
                lod50_values.append(lod50_val)
            if not math.isnan(lod95_val):
                lod95_values.append(lod95_val)

        lod50_panel = min(lod50_values) if lod50_values else math.nan
        lod95_panel = min(lod95_values) if lod95_values else math.nan
        threshold_panel = min(threshold_values) if threshold_values else math.nan

        interpretation = (
            "All expected species found"
            if n_found == n_expected and n_expected > 0
            else "No expected species found"
            if n_found == 0
            else "Partial recovery of expected species"
        )
        performance_interpretation = build_interpretation(
            pd.Series(
                {
                    "sensitivity": panel_sensitivity,
                    "specificity": specificity,
                    "false_positive_rate": false_positive_rate,
                    "lod50_spike_n": lod50_panel,
                }
            )
        )
        reported_species = sorted(set(found_species))
        off_target_species = [x for x in reported_species if x not in expected_species]

        n_reported = len(reported_species)
        n_off_target = len(off_target_species)
        rows.append(
            {
                "Workflow": workflow,
                "Metric": metric,
                "Run": run_name,
                "Expected species": "; ".join(expected_species),
                "Found species": "; ".join(found_species),
                "Missing expected species": "; ".join(missing_species),
                "N expected species": n_expected,
                "N reported species": n_reported,
                "N off-target species": n_off_target,
                "N found species": n_found,
                "Panel species recovery": panel_sensitivity,
                "Threshold": threshold_panel,
                "Specificity": specificity,
                "False positive rate": false_positive_rate,
                "LOD50": lod50_panel,
                "LOD95": lod95_panel,
                "Interpretation": interpretation,
                "Performance interpretation": performance_interpretation,
                "Per-species summary": " | ".join(per_species_summaries),
            }
        )

    panel_df = pd.DataFrame(rows)
    if panel_df.empty:
        return panel_df

    panel_df["lod50_rank"] = panel_df["LOD50"].fillna(10**12)
    panel_df["lod95_rank"] = panel_df["LOD95"].fillna(10**12)
    panel_df = panel_df.sort_values(
        by=[
            "Panel species recovery",
            "Specificity",
            "False positive rate",
            "lod50_rank",
            "lod95_rank",
            "N found species",
        ],
        ascending=[False, False, True, True, True, False],
    ).reset_index(drop=True)
    panel_df["Rank"] = range(1, len(panel_df) + 1)
    panel_df = panel_df.drop(columns=["lod50_rank", "lod95_rank"])
    return panel_df[
        [
            "Rank",
            "Workflow",
            "Metric",
            "Run",
            "Expected species",
            "Found species",
            "Missing expected species",
            "N expected species",
            "N found species",
            "N reported species",
            "N off-target species",
            "Panel species recovery",
            "Threshold",
            "Specificity",
            "False positive rate",
            "LOD50",
            "LOD95",
            "Interpretation",
            "Performance interpretation",
            "Per-species summary",
        ]
    ]



def build_reported_species_summary(
    detection_df: pd.DataFrame,
    threshold_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Build a verbose run-level table of all reported species and off-target hits.

    Parameters
    ----------
    detection_df : pd.DataFrame
        Detection table with thresholded calls.
    threshold_df : pd.DataFrame
        Per-target threshold table.

    Returns
    -------
    pd.DataFrame
        Verbose reported/off-target species summary.
    """
    if detection_df.empty:
        return pd.DataFrame()

    rows: list[dict[str, object]] = []
    grouped = detection_df.groupby(["workflow", "metric", "run_name"], dropna=False)

    for (workflow, metric, run_name), run_df in grouped:
        positive_df = run_df.loc[run_df["is_positive"]].copy()
        expected_species = sorted(pd.unique(positive_df["target_label"].astype(str)))

        all_species = sorted(pd.unique(run_df["target_label"].astype(str)))
        reported_any = sorted(
            pd.unique(
                run_df.loc[run_df["detected"] == True, "target_label"].astype(str)
            )
        )
        reported_positive = sorted(
            pd.unique(
                run_df.loc[
                    (run_df["is_positive"]) & (run_df["detected"] == True),
                    "target_label",
                ].astype(str)
            )
        )
        reported_negative = sorted(
            pd.unique(
                run_df.loc[
                    (run_df["is_negative"]) & (run_df["detected"] == True),
                    "target_label",
                ].astype(str)
            )
        )

        off_target_any = [x for x in reported_any if x not in expected_species]
        off_target_positive = [x for x in reported_positive if x not in expected_species]
        off_target_negative = [x for x in reported_negative if x not in expected_species]
        missing_expected = [x for x in expected_species if x not in reported_positive]

        threshold_rows = threshold_df.loc[
            (threshold_df["workflow"] == workflow)
            & (threshold_df["metric"] == metric)
        ].copy()
        threshold_values = pd.to_numeric(
            threshold_rows["threshold_value"], errors="coerce"
        ).dropna()
        threshold_panel = (
            float(threshold_values.min()) if not threshold_values.empty else math.nan
        )

        species_detail: list[str] = []
        for species in all_species:
            species_df = run_df.loc[run_df["target_label"] == species].copy()
            n_observations = int(species_df.shape[0])
            n_detected = int((species_df["detected"] == True).sum())
            pos_detected = int(
                ((species_df["is_positive"]) & (species_df["detected"] == True)).sum()
            )
            neg_detected = int(
                ((species_df["is_negative"]) & (species_df["detected"] == True)).sum()
            )
            detection_rate = safe_divide(n_detected, n_observations)
            species_detail.append(
                f"{species}: detected={n_detected}/{n_observations}, "
                f"positive={pos_detected}, negative={neg_detected}, "
                f"rate={format_value(detection_rate)}"
            )

        if reported_any and not off_target_any and not missing_expected:
            interpretation = "Only expected tracked species reported"
        elif reported_any and off_target_any:
            interpretation = "Off-target species also reported"
        elif not reported_any:
            interpretation = "No species reported above threshold"
        else:
            interpretation = "Partial expected-species recovery"

        rows.append(
            {
                "Workflow": workflow,
                "Metric": metric,
                "Run": run_name,
                "Expected species": "; ".join(expected_species),
                "All reported species": "; ".join(reported_any),
                "Reported species in positives": "; ".join(reported_positive),
                "Reported species in negatives": "; ".join(reported_negative),
                "Off-target species": "; ".join(off_target_any),
                "Off-target species in positives": "; ".join(off_target_positive),
                "Off-target species in negatives": "; ".join(off_target_negative),
                "Missing expected species": "; ".join(missing_expected),
                "N expected species": len(expected_species),
                "N reported species": len(reported_any),
                "N off-target species": len(off_target_any),
                "Threshold": threshold_panel,
                "Interpretation": interpretation,
                "Reported species detail": " | ".join(species_detail),
                "Note": (
                    "This table is limited to species labels present in combined_long.tsv. "
                    "Methods summarised only as single_target cannot expose off-target taxa here."
                ),
            }
        )

    reported_df = pd.DataFrame(rows)
    if reported_df.empty:
        return reported_df

    reported_df = reported_df.sort_values(
        by=["N off-target species", "N reported species", "Workflow", "Metric", "Run"],
        ascending=[True, False, True, True, True],
    ).reset_index(drop=True)
    reported_df.insert(0, "Rank", range(1, len(reported_df) + 1))
    return reported_df


def format_value(value: object) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return f"{value:.3f}"
    return str(value)


def value_to_css_class(column: str, value: object) -> str:
    if pd.isna(value):
        return "cell-empty"

    good_columns = {
        "Sensitivity",
        "Specificity",
        "Precision",
        "F1",
        "Panel species recovery",
        "balanced_accuracy",
        "accuracy",
    }
    bad_columns = {
        "False positive rate",
        "False negative rate",
        "false_positive_rate",
        "false_negative_rate",
    }

    if not isinstance(value, (int, float)):
        return ""

    if column in good_columns:
        if value >= 0.90:
            return "cell-good"
        if value >= 0.50:
            return "cell-mid"
        return "cell-bad"

    if column in bad_columns:
        if value <= 0.10:
            return "cell-good"
        if value <= 0.50:
            return "cell-mid"
        return "cell-bad"

    return ""


def dataframe_to_html_table(
    dataframe: pd.DataFrame,
    table_class: str = "data-table",
    compact: bool = False,
) -> str:
    if dataframe.empty:
        return '<p class="muted">No rows available.</p>'

    header_cells = "".join(
        f"<th>{html.escape(str(column))}</th>"
        for column in dataframe.columns
    )

    body_rows = []
    for _, row in dataframe.iterrows():
        cells = []
        for column, value in row.items():
            css_class = value_to_css_class(str(column), value)
            extra_class = f' class="{css_class}"' if css_class else ""
            cells.append(f"<td{extra_class}>{html.escape(format_value(value))}</td>")
        body_rows.append(f"<tr>{''.join(cells)}</tr>")

    wrapper_class = "table-wrap compact-wrap" if compact else "table-wrap"
    return (
        f'<div class="{wrapper_class}">'
        f"<table class=\"{html.escape(table_class)}\">"
        f"<thead><tr>{header_cells}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody>"
        f"</table>"
        f"</div>"
    )


def build_html_page(
    compact_df: pd.DataFrame,
    panel_species_df: pd.DataFrame,
    reported_species_df: pd.DataFrame,
    method_df: pd.DataFrame,
    by_spike_df: pd.DataFrame,
    definitions_df: pd.DataFrame,
    plot_explanations_df: pd.DataFrame,
) -> str:
    style = """
    <style>
      body {
        font-family: Arial, Helvetica, sans-serif;
        margin: 0;
        padding: 0;
        color: #1a1a1a;
        background: #f7f9fc;
      }
      .container {
        max-width: 1520px;
        margin: 0 auto;
        padding: 32px;
        background: #ffffff;
      }
      h1, h2, h3 {
        color: #1f4e79;
      }
      .table-wrap {
        overflow-x: auto;
        border: 1px solid #dbe5f0;
        border-radius: 10px;
        margin-top: 14px;
        background: #ffffff;
      }
      .compact-wrap {
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.05);
      }
      .data-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 14px;
      }
      .data-table thead th {
        position: sticky;
        top: 0;
        background: #1f4e79;
        color: white;
        text-align: left;
        padding: 10px 8px;
        white-space: nowrap;
        z-index: 2;
      }
      .data-table td {
        border-bottom: 1px solid #e6edf5;
        padding: 8px;
        vertical-align: top;
        white-space: nowrap;
      }
      .data-table tbody tr:nth-child(even) {
        background: #fbfdff;
      }
      .muted {
        color: #666666;
      }
      .section-note {
        margin-top: 8px;
        color: #333333;
      }
      .small-note {
        font-size: 13px;
        color: #555555;
      }
      .cell-good {
        background: #e9f7ef;
        font-weight: 600;
      }
      .cell-mid {
        background: #fff8e1;
      }
      .cell-bad {
        background: #fdecea;
      }
      .cell-empty {
        color: #999999;
      }
      details {
        margin-top: 18px;
      }
      summary {
        cursor: pointer;
        font-weight: 600;
        color: #1f4e79;
      }
    </style>
    """

    compact_table = dataframe_to_html_table(compact_df, compact=True)
    panel_species_table = dataframe_to_html_table(panel_species_df, compact=True)
    reported_species_table = dataframe_to_html_table(reported_species_df, compact=True)
    method_table = dataframe_to_html_table(method_df, compact=False)
    by_spike_table = dataframe_to_html_table(by_spike_df, compact=False)
    definitions_table = dataframe_to_html_table(definitions_df, compact=False)
    plot_table = dataframe_to_html_table(plot_explanations_df, compact=False)

    return f"""<!DOCTYPE html>
<html lang=\"en\">
<head>
  <meta charset=\"utf-8\">
  <title>Method performance summary</title>
  {style}
</head>
<body>
  <div class=\"container\">
    <h1>Method performance summary</h1>
    <p class=\"muted\">
      Positive observations are non-shuffled rows with spike_n greater than 0.
      Negative observations are shuffled-control rows or rows with spike_n equal
      to 0. Detection is called using the selected threshold rule.
    </p>

    <h2>Headline comparison table</h2>
    <p class=\"section-note\">
      This compact table is intended for rapid comparison across methods.
      Methods are ranked primarily by balanced accuracy, then specificity,
      then false-positive rate, then sensitivity, then LOD.
    </p>
    {compact_table}

    <h2>Panel species summary</h2>
    <p class=\"section-note\">
      This table shows the full set of species expected in each run panel and
      the set of tracked target species that were actually found under the
      chosen threshold rule.
    </p>
    {panel_species_table}

    <details open>
      <summary>Show full method-level performance table</summary>
      <p class=\"small-note\">
        This detailed table includes confusion counts, threshold statistics,
        and the full set of derived performance metrics.
      </p>
      {method_table}
    </details>

    <details>
      <summary>Show detection rate by spike level</summary>
      <p class=\"small-note\">
        This table shows how often each method is called positive at each spike
        level among positive observations.
      </p>
      {by_spike_table}
    </details>

    <h2>Definitions of reported performance statistics</h2>
    {definitions_table}

    <h2>How to interpret the plots in the HTML report</h2>
    {plot_table}
  </div>
</body>
</html>
"""


def build_html_fragment(
    compact_df: pd.DataFrame,
    panel_species_df: pd.DataFrame,
    reported_species_df: pd.DataFrame,
    method_df: pd.DataFrame,
    definitions_df: pd.DataFrame,
    plot_explanations_df: pd.DataFrame,
) -> str:
    compact_table = dataframe_to_html_table(compact_df, compact=True)
    panel_species_table = dataframe_to_html_table(panel_species_df, compact=True)
    reported_species_table = dataframe_to_html_table(reported_species_df, compact=True)
    method_table = dataframe_to_html_table(method_df, compact=False)
    definitions_table = dataframe_to_html_table(definitions_df, compact=False)
    plot_table = dataframe_to_html_table(plot_explanations_df, compact=False)

    return f"""
<section class=\"report-section\">
  <h2>Method performance summary</h2>
  <p class=\"muted\">
    Sensitivity, recall, and true positive rate are equivalent here and are
    all reported for convenience. Thresholding is method-specific and based on
    the selected negative-control rule.
  </p>

  <h3>Headline comparison table</h3>
  {compact_table}

  <h3>Panel species summary</h3>
  {panel_species_table}

  <h3>All reported species / off-target species</h3>
  {reported_species_table}

  <details>
    <summary>Show full detail table</summary>
    {method_table}
  </details>

  <h3>Definitions of reported statistics</h3>
  {definitions_table}

  <h3>How to interpret the plots</h3>
  {plot_table}
</section>
"""


def format_excel_sheet(
    workbook_path: Path,
    percentage_columns: set[str],
    integer_columns: set[str],
) -> None:
    workbook = load_workbook(filename=workbook_path)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(vertical="center", wrap_text=True)

    for worksheet in workbook.worksheets:
        if worksheet.max_row == 0 or worksheet.max_column == 0:
            continue

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        headers = [cell.value for cell in worksheet[1]]

        for idx, header in enumerate(headers, start=1):
            if header is None:
                continue
            column_letter = get_column_letter(idx)
            max_len = len(str(header))

            for row in worksheet.iter_rows(
                min_row=2,
                min_col=idx,
                max_col=idx,
                values_only=False,
            ):
                cell = row[0]
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))

                    if str(header) in percentage_columns and isinstance(cell.value, (int, float)):
                        cell.number_format = "0.000"
                    elif str(header) in integer_columns and isinstance(cell.value, (int, float)):
                        cell.number_format = "0"

            worksheet.column_dimensions[column_letter].width = min(max_len + 2, 42)

    workbook.save(filename=workbook_path)


def main() -> None:
    args = parse_args()

    out_dir = Path(args.out_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    combined_long = pd.read_csv(args.combined_long_tsv, sep="\t")
    combined_long = normalise_combined_long(combined_long)

    detection_df, threshold_df = compute_detection_calls(
        dataframe=combined_long,
        threshold_mode=args.threshold_mode,
        min_detect_value=args.min_detect_value,
        sd_multiplier=args.sd_multiplier,
    )

    method_df = summarise_method_performance(detection_df, threshold_df)
    compact_df = build_compact_table(method_df)
    panel_species_df = build_panel_species_summary(detection_df, threshold_df)
    by_spike_df = summarise_detection_by_spike(detection_df)
    definitions_df = pd.DataFrame(METRIC_DEFINITIONS)
    plot_explanations_df = pd.DataFrame(REPORT_PLOT_EXPLANATIONS)

    method_tsv = out_dir / "method_performance.tsv"
    compact_tsv = out_dir / "method_performance_compact.tsv"
    panel_species_tsv = out_dir / "method_panel_species_summary.tsv"
    by_spike_tsv = out_dir / "method_performance_by_spike.tsv"
    definitions_tsv = out_dir / "metric_definitions.tsv"
    xlsx_path = out_dir / "method_performance.xlsx"
    page_html = out_dir / "method_performance.html"
    fragment_html = out_dir / "method_performance_fragment.html"

    method_df.to_csv(method_tsv, sep="\t", index=False)
    compact_df.to_csv(compact_tsv, sep="\t", index=False)
    panel_species_df.to_csv(panel_species_tsv, sep="\t", index=False)
    reported_species_df.to_csv(reported_species_tsv, sep="\t", index=False)
    by_spike_df.to_csv(by_spike_tsv, sep="\t", index=False)
    definitions_df.to_csv(definitions_tsv, sep="\t", index=False)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        compact_df.to_excel(writer, sheet_name="headline_table", index=False)
        panel_species_df.to_excel(writer, sheet_name="panel_species_summary", index=False)
        reported_species_df.to_excel(writer, sheet_name="reported_offtargets", index=False)
        method_df.to_excel(writer, sheet_name="full_method_table", index=False)
        by_spike_df.to_excel(writer, sheet_name="by_spike", index=False)
        definitions_df.to_excel(writer, sheet_name="definitions", index=False)
        plot_explanations_df.to_excel(writer, sheet_name="plot_interpretation", index=False)

    format_excel_sheet(
        workbook_path=xlsx_path,
        percentage_columns={
            "Sensitivity",
            "Specificity",
            "Precision",
            "F1",
            "False positive rate",
            "False negative rate",
            "Panel species recovery",
            "sensitivity",
            "recall",
            "true_positive_rate",
            "specificity",
            "false_positive_rate",
            "false_negative_rate",
            "precision",
            "negative_predictive_value",
            "accuracy",
            "balanced_accuracy",
            "f1_score",
            "youden_j",
            "detection_rate",
        },
        integer_columns={
            "Rank",
            "N expected species",
            "N found species",
            "N reported species",
            "N off-target species",
            "TP",
            "FN",
            "FP",
            "TN",
            "LOD50",
            "LOD95",
            "tp",
            "fn",
            "fp",
            "tn",
            "n_positive",
            "n_negative",
            "lod50_spike_n",
            "lod95_spike_n",
            "spike_n",
            "n_observations",
            "n_detected",
        },
    )

    page_html.write_text(
        build_html_page(
            compact_df=compact_df,
            panel_species_df=panel_species_df,
            reported_species_df=reported_species_df,
            method_df=method_df,
            by_spike_df=by_spike_df,
            definitions_df=definitions_df,
            plot_explanations_df=plot_explanations_df,
        ),
        encoding="utf-8",
    )
    fragment_html.write_text(
        build_html_fragment(
            compact_df=compact_df,
            panel_species_df=panel_species_df,
            reported_species_df=reported_species_df,
            method_df=method_df,
            definitions_df=definitions_df,
            plot_explanations_df=plot_explanations_df,
        ),
        encoding="utf-8",
    )

    print(f"[INFO] Wrote: {method_tsv}")
    print(f"[INFO] Wrote: {compact_tsv}")
    print(f"[INFO] Wrote: {panel_species_tsv}")
    print(f"[INFO] Wrote: {reported_species_tsv}")
    print(f"[INFO] Wrote: {by_spike_tsv}")
    print(f"[INFO] Wrote: {definitions_tsv}")
    print(f"[INFO] Wrote: {xlsx_path}")
    print(f"[INFO] Wrote: {page_html}")
    print(f"[INFO] Wrote: {fragment_html}")


if __name__ == "__main__":
    main()
