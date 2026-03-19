#!/usr/bin/env python3
"""Summarise ONT spike-in run outputs, plots, and Excel workbook outputs.

This script recursively scans one or more run directories produced by the
spike-in shell workflows and collates their summary TSV files into a harmonised
set of tables. It is designed to be fault tolerant: missing files, partial
runs, empty tables, or malformed rows are recorded as warnings rather than
causing the script to terminate.

Outputs
-------
The script writes:
- run_manifest.tsv
- combined_wide.tsv
- combined_long.tsv
- missing_or_problematic_files.tsv
- plot_manifest.tsv
- summary_statistics.tsv
- spikein_summary.xlsx
- plots/*.png and plots/*.pdf
"""

from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SUMMARY_FILE_MAP = {
    "spikein_summary.tsv": "single_read",
    "spikein_flye_summary.tsv": "single_assembly",
    "spikein_multi_summary.tsv": "multi_read",
    "spikein_multi_flye_summary.tsv": "multi_assembly",
}

READ_METRIC_COLUMNS = [
    "kraken_target_reads",
    "minimap_target_alignments",
]

ASSEMBLY_METRIC_COLUMNS = [
    "kraken_target_contigs",
]


@dataclass
class IssueRecord:
    """Store a warning or non-fatal problem encountered during parsing."""

    category: str
    path: str
    detail: str


@dataclass
class PlotRecord:
    """Store metadata for a generated plot file."""

    plot_group: str
    workflow_type: str
    metric_name: str
    target_label: str
    png_path: str
    pdf_path: str
    n_points: int


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(vertical="center", wrap_text=True)


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments for the summarisation workflow."""
    parser = argparse.ArgumentParser(
        description=(
            "Recursively summarise ONT spike-in run outputs from one or more "
            "directories without failing on missing results."
        )
    )
    parser.add_argument(
        "--input_dirs",
        nargs="+",
        required=True,
        help="One or more run directories to scan recursively.",
    )
    parser.add_argument(
        "--out_dir",
        required=True,
        help="Output directory for combined tables, workbook, and plots.",
    )
    parser.add_argument(
        "--glob_patterns",
        nargs="*",
        default=list(SUMMARY_FILE_MAP.keys()),
        help="Optional summary file names to search for.",
    )
    parser.add_argument(
        "--plot_min_points",
        type=int,
        default=2,
        help="Minimum number of rows required before a plot is generated.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print progress messages during parsing and plotting.",
    )
    return parser.parse_args()



def log(message: str, verbose: bool) -> None:
    """Print a progress message when verbose mode is enabled."""
    if verbose:
        print(message)



def discover_summary_files(
    input_dirs: Iterable[str],
    glob_patterns: Iterable[str],
    verbose: bool,
) -> list[Path]:
    """Recursively discover recognised summary TSV files."""
    found: list[Path] = []
    seen: set[Path] = set()

    for input_dir in input_dirs:
        base = Path(input_dir).expanduser().resolve()
        if not base.exists():
            log(f"[WARN] Input directory does not exist: {base}", verbose=verbose)
            continue

        for pattern in glob_patterns:
            for path in base.rglob(pattern):
                if path.is_file() and path not in seen:
                    found.append(path)
                    seen.add(path)

    return sorted(found)



def read_key_value_tsv(path: Path) -> dict[str, str]:
    """Read a simple two-column parameter/value TSV into a dictionary."""
    mapping: dict[str, str] = {}
    if not path.exists() or path.stat().st_size == 0:
        return mapping

    try:
        dataframe = pd.read_csv(path, sep="\t", dtype=str)
    except Exception:
        return mapping

    if {"parameter", "value"}.issubset(dataframe.columns):
        for _, row in dataframe.iterrows():
            parameter = str(row.get("parameter", "")).strip()
            value = str(row.get("value", "")).strip()
            if parameter:
                mapping[parameter] = value

    return mapping



def detect_shuffle_context(run_dir: Path) -> tuple[bool, Optional[Path], Optional[Path]]:
    """Detect whether a run directory belongs to a shuffled control analysis."""
    run_metadata = run_dir / "run_metadata.tsv"
    shuffle_metadata = run_dir / "shuffle_metadata.tsv"

    if run_metadata.exists() or shuffle_metadata.exists():
        return True, run_metadata if run_metadata.exists() else None, (
            shuffle_metadata if shuffle_metadata.exists() else None
        )

    parent = run_dir.parent
    run_metadata = parent / "run_metadata.tsv"
    shuffle_metadata = parent / "shuffle_metadata.tsv"
    if run_metadata.exists() or shuffle_metadata.exists():
        return True, run_metadata if run_metadata.exists() else None, (
            shuffle_metadata if shuffle_metadata.exists() else None
        )

    return False, None, None



def safe_read_tsv(path: Path, issues: list[IssueRecord]) -> Optional[pd.DataFrame]:
    """Read a TSV file safely, recording problems rather than failing."""
    try:
        if not path.exists():
            issues.append(IssueRecord("missing_file", str(path), "File not found."))
            return None
        if path.stat().st_size == 0:
            issues.append(IssueRecord("empty_file", str(path), "File is empty."))
            return None
        return pd.read_csv(path, sep="\t", dtype=str)
    except Exception as exc:  # noqa: BLE001
        issues.append(IssueRecord("read_error", str(path), f"{type(exc).__name__}: {exc}"))
        return None



def coerce_numeric_columns(dataframe: pd.DataFrame, columns: Iterable[str]) -> pd.DataFrame:
    """Convert selected columns to numeric when present."""
    for column in columns:
        if column in dataframe.columns:
            dataframe[column] = pd.to_numeric(dataframe[column], errors="coerce")
    return dataframe



def add_common_metadata(
    dataframe: pd.DataFrame,
    summary_path: Path,
    workflow_type: str,
    is_shuffled_control: bool,
    run_metadata_path: Optional[Path],
    shuffle_metadata_path: Optional[Path],
) -> pd.DataFrame:
    """Append common metadata fields to a parsed summary table."""
    run_dir = summary_path.parent
    dataframe = dataframe.copy()
    dataframe["workflow_type"] = workflow_type
    dataframe["summary_path"] = str(summary_path)
    dataframe["run_dir"] = str(run_dir)
    dataframe["run_name"] = run_dir.name
    dataframe["is_shuffled_control"] = is_shuffled_control
    dataframe["run_metadata_path"] = (
        str(run_metadata_path) if run_metadata_path is not None else pd.NA
    )
    dataframe["shuffle_metadata_path"] = (
        str(shuffle_metadata_path) if shuffle_metadata_path is not None else pd.NA
    )
    dataframe["source_file_name"] = summary_path.name

    metadata_map: dict[str, str] = {}
    if run_metadata_path is not None:
        metadata_map.update(read_key_value_tsv(run_metadata_path))

    dataframe["metadata_target_label"] = metadata_map.get("target_label", pd.NA)
    dataframe["metadata_pathogen_fasta"] = metadata_map.get("pathogen_fasta", pd.NA)
    dataframe["metadata_shuffled_fasta"] = metadata_map.get("shuffled_fasta", pd.NA)
    return dataframe



def standardise_single_read(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Standardise single-genome read-level summary columns."""
    dataframe = dataframe.copy()
    dataframe["spike_n_per_genome"] = dataframe.get("spike_n")
    dataframe["total_spike_n"] = dataframe.get("spike_n")
    dataframe["n_genomes"] = 1
    dataframe["target_label"] = dataframe.get("target_label", dataframe.get("metadata_target_label"))
    dataframe["target_group"] = "single"
    dataframe["assembly_n_contigs"] = pd.NA
    dataframe["assembly_total_bases"] = pd.NA
    dataframe["kraken_target_contigs"] = pd.NA
    dataframe["kraken_target_reads_g1"] = dataframe.get("kraken_target_reads")
    dataframe["minimap_target_alignments_g1"] = dataframe.get(
        "minimap_target_alignments"
    )
    dataframe["kraken_target_contigs_g1"] = pd.NA
    if "kraken_classified_contigs" not in dataframe.columns:
        dataframe["kraken_classified_contigs"] = pd.NA
    return dataframe



def standardise_single_assembly(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Standardise single-genome assembly summary columns."""
    dataframe = dataframe.copy()
    dataframe["spike_n_per_genome"] = dataframe.get("spike_n")
    dataframe["total_spike_n"] = dataframe.get("spike_n")
    dataframe["n_genomes"] = 1
    dataframe["target_label"] = dataframe.get("target_label", dataframe.get("metadata_target_label"))
    dataframe["target_group"] = "single"
    dataframe["kraken_target_reads"] = pd.NA
    dataframe["minimap_target_alignments"] = pd.NA
    dataframe["kraken_target_reads_g1"] = pd.NA
    dataframe["minimap_target_alignments_g1"] = pd.NA
    dataframe["kraken_target_contigs_g1"] = dataframe.get("kraken_target_contigs")
    if "kraken_classified_reads" not in dataframe.columns:
        dataframe["kraken_classified_reads"] = pd.NA
    return dataframe



def standardise_multi_read(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Standardise multi-genome read-level summary columns."""
    dataframe = dataframe.copy()
    dataframe["target_group"] = "multi"
    dataframe["kraken_target_reads"] = pd.NA
    dataframe["minimap_target_alignments"] = pd.NA
    dataframe["kraken_target_contigs"] = pd.NA
    dataframe["assembly_n_contigs"] = pd.NA
    dataframe["assembly_total_bases"] = pd.NA
    if "kraken_classified_contigs" not in dataframe.columns:
        dataframe["kraken_classified_contigs"] = pd.NA
    return dataframe



def standardise_multi_assembly(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Standardise multi-genome assembly summary columns."""
    dataframe = dataframe.copy()
    dataframe["target_group"] = "multi"
    dataframe["kraken_target_reads"] = pd.NA
    dataframe["minimap_target_alignments"] = pd.NA
    if "kraken_classified_reads" not in dataframe.columns:
        dataframe["kraken_classified_reads"] = pd.NA
    return dataframe



def infer_multi_genome_count(dataframe: pd.DataFrame) -> pd.Series:
    """Infer the number of genomes represented in a multi-genome table."""
    label_columns = [
        column for column in dataframe.columns
        if re.match(r"^target_label_g\d+$", str(column))
    ]
    if "n_genomes" in dataframe.columns:
        return dataframe["n_genomes"]
    if not label_columns:
        return pd.Series([pd.NA] * len(dataframe), index=dataframe.index)
    counts = dataframe[label_columns].notna().sum(axis=1)
    return counts.replace(0, pd.NA)



def parse_summary_file(
    path: Path,
    issues: list[IssueRecord],
    verbose: bool,
) -> Optional[pd.DataFrame]:
    """Parse a recognised summary TSV into a standardised wide table."""
    workflow_type = SUMMARY_FILE_MAP.get(path.name)
    if workflow_type is None:
        issues.append(IssueRecord("unknown_summary_type", str(path), "Unrecognised file name."))
        return None

    dataframe = safe_read_tsv(path=path, issues=issues)
    if dataframe is None:
        return None
    if dataframe.empty:
        issues.append(IssueRecord("empty_table", str(path), "Summary table has no rows."))
        return None

    is_shuffled, run_meta_path, shuffle_meta_path = detect_shuffle_context(run_dir=path.parent)
    dataframe = add_common_metadata(
        dataframe=dataframe,
        summary_path=path,
        workflow_type=workflow_type,
        is_shuffled_control=is_shuffled,
        run_metadata_path=run_meta_path,
        shuffle_metadata_path=shuffle_meta_path,
    )

    if workflow_type == "single_read":
        dataframe = standardise_single_read(dataframe=dataframe)
    elif workflow_type == "single_assembly":
        dataframe = standardise_single_assembly(dataframe=dataframe)
    elif workflow_type == "multi_read":
        dataframe = standardise_multi_read(dataframe=dataframe)
    elif workflow_type == "multi_assembly":
        dataframe = standardise_multi_assembly(dataframe=dataframe)

    if workflow_type.startswith("multi"):
        if "spike_n_per_genome" not in dataframe.columns:
            dataframe["spike_n_per_genome"] = dataframe.get("spike_n")
        if "n_genomes" not in dataframe.columns:
            dataframe["n_genomes"] = infer_multi_genome_count(dataframe)
        if "total_spike_n" not in dataframe.columns:
            try:
                dataframe["total_spike_n"] = (
                    pd.to_numeric(dataframe["spike_n_per_genome"], errors="coerce")
                    * pd.to_numeric(dataframe["n_genomes"], errors="coerce")
                )
            except Exception:  # noqa: BLE001
                dataframe["total_spike_n"] = pd.NA

    numeric_columns = [
        "replicate",
        "spike_n",
        "spike_n_per_genome",
        "n_genomes",
        "total_spike_n",
        "kraken_classified_reads",
        "kraken_classified_contigs",
        "kraken_target_reads",
        "minimap_target_alignments",
        "kraken_target_contigs",
        "assembly_n_contigs",
        "assembly_total_bases",
    ]
    dynamic_numeric = [
        column for column in dataframe.columns
        if re.match(r"^(kraken_target_reads|minimap_target_alignments|kraken_target_contigs)_g\d+$", column)
    ]
    dataframe = coerce_numeric_columns(
        dataframe=dataframe,
        columns=numeric_columns + dynamic_numeric,
    )

    log(f"[INFO] Parsed {path}", verbose=verbose)
    return dataframe



def build_run_manifest(dataframes: list[pd.DataFrame]) -> pd.DataFrame:
    """Build a one-row-per-run manifest from parsed summary tables."""
    records: list[dict[str, object]] = []
    for dataframe in dataframes:
        if dataframe.empty:
            continue
        first = dataframe.iloc[0]
        records.append(
            {
                "workflow_type": first.get("workflow_type"),
                "run_name": first.get("run_name"),
                "run_dir": first.get("run_dir"),
                "summary_path": first.get("summary_path"),
                "source_file_name": first.get("source_file_name"),
                "is_shuffled_control": first.get("is_shuffled_control"),
                "target_label": first.get("target_label"),
                "n_rows": len(dataframe),
            }
        )
    if not records:
        return pd.DataFrame(
            columns=[
                "workflow_type",
                "run_name",
                "run_dir",
                "summary_path",
                "source_file_name",
                "is_shuffled_control",
                "target_label",
                "n_rows",
            ]
        )
    return pd.DataFrame.from_records(records)



def extract_multi_label_map(row: pd.Series) -> dict[str, str]:
    """Extract genome-index to target-label mappings from a multi-genome row."""
    mapping: dict[str, str] = {}
    for column, value in row.items():
        match = re.match(r"^target_label_g(\d+)$", str(column))
        if match and pd.notna(value):
            mapping[match.group(1)] = str(value)
    return mapping



def wide_to_long(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Convert the standardised wide table to a long plotting table."""
    records: list[dict[str, object]] = []

    for _, row in dataframe.iterrows():
        common = {
            "workflow_type": row.get("workflow_type"),
            "run_name": row.get("run_name"),
            "run_dir": row.get("run_dir"),
            "summary_path": row.get("summary_path"),
            "is_shuffled_control": row.get("is_shuffled_control"),
            "replicate": row.get("replicate"),
            "spike_n_per_genome": row.get("spike_n_per_genome"),
            "total_spike_n": row.get("total_spike_n"),
            "n_genomes": row.get("n_genomes"),
            "assembly_n_contigs": row.get("assembly_n_contigs"),
            "assembly_total_bases": row.get("assembly_total_bases"),
        }

        if row.get("workflow_type") == "single_read":
            for metric_name in READ_METRIC_COLUMNS:
                value = row.get(metric_name)
                if pd.notna(value):
                    records.append(
                        {
                            **common,
                            "target_label": row.get("target_label")
                            if pd.notna(row.get("target_label"))
                            else "single_target",
                            "metric_name": metric_name,
                            "metric_value": value,
                        }
                    )
        elif row.get("workflow_type") == "single_assembly":
            for metric_name in ASSEMBLY_METRIC_COLUMNS:
                value = row.get(metric_name)
                if pd.notna(value):
                    records.append(
                        {
                            **common,
                            "target_label": row.get("target_label")
                            if pd.notna(row.get("target_label"))
                            else "single_target",
                            "metric_name": metric_name,
                            "metric_value": value,
                        }
                    )
        elif row.get("workflow_type") in {"multi_read", "multi_assembly"}:
            label_map = extract_multi_label_map(row=row)
            for genome_idx, target_label in label_map.items():
                if row.get("workflow_type") == "multi_read":
                    metric_columns = [
                        f"kraken_target_reads_g{genome_idx}",
                        f"minimap_target_alignments_g{genome_idx}",
                    ]
                else:
                    metric_columns = [f"kraken_target_contigs_g{genome_idx}"]

                for metric_column in metric_columns:
                    value = row.get(metric_column)
                    if pd.notna(value):
                        records.append(
                            {
                                **common,
                                "target_label": target_label,
                                "metric_name": re.sub(r"_g\d+$", "", metric_column),
                                "metric_value": value,
                            }
                        )

    long_df = pd.DataFrame.from_records(records)
    if long_df.empty:
        return pd.DataFrame(
            columns=[
                "workflow_type",
                "run_name",
                "run_dir",
                "summary_path",
                "is_shuffled_control",
                "replicate",
                "spike_n_per_genome",
                "total_spike_n",
                "n_genomes",
                "assembly_n_contigs",
                "assembly_total_bases",
                "target_label",
                "metric_name",
                "metric_value",
            ]
        )

    return coerce_numeric_columns(
        dataframe=long_df,
        columns=[
            "replicate",
            "spike_n_per_genome",
            "total_spike_n",
            "n_genomes",
            "assembly_n_contigs",
            "assembly_total_bases",
            "metric_value",
        ],
    )



def sanitise_filename(value: str) -> str:
    """Convert a label into a filesystem-safe stem."""
    value = value.strip().replace(" ", "_")
    value = re.sub(r"[^A-Za-z0-9_.-]+", "_", value)
    value = re.sub(r"_+", "_", value)
    return value.strip("_") or "plot"



def maybe_log_scale(values: pd.Series) -> bool:
    """Return True when spike values span a wide enough range for log scaling."""
    positive = values[pd.notna(values) & (values > 0)]
    if positive.empty:
        return False
    return positive.max() / max(positive.min(), 1) >= 50



def generate_metric_plot(
    plot_df: pd.DataFrame,
    out_dir: Path,
    plot_group: str,
    workflow_type: str,
    metric_name: str,
    target_label: str,
    issues: list[IssueRecord],
    y_label: str,
    title_suffix: str,
) -> Optional[PlotRecord]:
    """Generate one plot for a workflow-metric-target combination."""
    if plot_df.empty:
        return None

    plot_df = plot_df.sort_values(
        by=["total_spike_n", "spike_n_per_genome", "replicate"],
        kind="stable",
    )

    fig, ax = plt.subplots(figsize=(7, 5))

    for run_name, run_df in plot_df.groupby("run_name", dropna=False):
        run_df = run_df.sort_values(by=["total_spike_n", "spike_n_per_genome"])
        grouped = run_df.groupby("total_spike_n", dropna=False, as_index=False)["metric_value"].median()
        x_vals = grouped["total_spike_n"]
        y_vals = grouped["metric_value"]
        label = str(run_name)
        if bool(run_df["is_shuffled_control"].fillna(False).iloc[0]):
            label = f"{label} (shuffled)"
        ax.plot(x_vals, y_vals, marker="o", linestyle="-", label=label)

    if maybe_log_scale(values=plot_df["total_spike_n"]):
        positive_min = plot_df.loc[plot_df["total_spike_n"] > 0, "total_spike_n"].min()
        if pd.notna(positive_min):
            ax.set_xscale("symlog", linthresh=max(1, positive_min))

    ax.set_xlabel("Total spiked reads")
    ax.set_ylabel(y_label)
    ax.set_title(f"{workflow_type}: {target_label} - {title_suffix}")
    ax.legend(loc="best", fontsize=8)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    stem = sanitise_filename(f"{plot_group}__{workflow_type}__{target_label}__{metric_name}")
    png_path = out_dir / f"{stem}.png"
    pdf_path = out_dir / f"{stem}.pdf"

    try:
        fig.savefig(png_path, dpi=300, bbox_inches="tight")
        fig.savefig(pdf_path, bbox_inches="tight")
    except Exception as exc:  # noqa: BLE001
        issues.append(IssueRecord("plot_error", str(png_path), f"{type(exc).__name__}: {exc}"))
        plt.close(fig)
        return None

    plt.close(fig)
    return PlotRecord(
        plot_group=plot_group,
        workflow_type=workflow_type,
        metric_name=metric_name,
        target_label=target_label,
        png_path=str(png_path),
        pdf_path=str(pdf_path),
        n_points=len(plot_df),
    )



def build_baseline_adjusted_table(long_df: pd.DataFrame) -> pd.DataFrame:
    """Create a baseline-adjusted metric table using spike_n == 0 as baseline."""
    if long_df.empty:
        return long_df.copy()

    baseline_keys = ["workflow_type", "run_name", "target_label", "metric_name"]
    baseline_df = (
        long_df.loc[long_df["total_spike_n"] == 0, baseline_keys + ["metric_value"]]
        .rename(columns={"metric_value": "baseline_value"})
        .drop_duplicates(subset=baseline_keys)
    )
    adjusted_df = long_df.merge(baseline_df, on=baseline_keys, how="left")
    adjusted_df["metric_value"] = adjusted_df["metric_value"] - adjusted_df["baseline_value"].fillna(0)
    return adjusted_df.drop(columns=["baseline_value"])



def build_efficiency_table(long_df: pd.DataFrame) -> pd.DataFrame:
    """Create a detection-efficiency table using metric over total spiked reads."""
    if long_df.empty:
        return long_df.copy()

    efficiency_df = long_df.copy()
    denominator = pd.to_numeric(efficiency_df["total_spike_n"], errors="coerce")
    efficiency_df = efficiency_df.loc[denominator > 0].copy()
    denominator = pd.to_numeric(efficiency_df["total_spike_n"], errors="coerce")
    efficiency_df["metric_value"] = pd.to_numeric(
        efficiency_df["metric_value"], errors="coerce"
    ) / denominator
    return efficiency_df



def build_completion_table(run_manifest: pd.DataFrame, issues_df: pd.DataFrame) -> pd.DataFrame:
    """Build a simple completion-status table for plotting."""
    n_runs = len(run_manifest)
    n_issue_paths = issues_df["path"].nunique() if not issues_df.empty else 0
    n_problem_records = len(issues_df)
    return pd.DataFrame(
        {
            "status": ["runs_found", "issue_paths", "issue_records"],
            "count": [n_runs, n_issue_paths, n_problem_records],
        }
    )



def generate_completion_plot(
    completion_df: pd.DataFrame,
    plots_dir: Path,
    issues: list[IssueRecord],
) -> Optional[PlotRecord]:
    """Generate a simple completion-status bar plot."""
    if completion_df.empty:
        return None

    fig, ax = plt.subplots(figsize=(6, 4))
    ax.bar(completion_df["status"], completion_df["count"])
    ax.set_ylabel("Count")
    ax.set_title("Run discovery and issue summary")
    ax.grid(True, axis="y", alpha=0.3)
    fig.tight_layout()

    png_path = plots_dir / "qc__completion_summary.png"
    pdf_path = plots_dir / "qc__completion_summary.pdf"
    try:
        fig.savefig(png_path, dpi=300, bbox_inches="tight")
        fig.savefig(pdf_path, bbox_inches="tight")
    except Exception as exc:  # noqa: BLE001
        issues.append(IssueRecord("plot_error", str(png_path), f"{type(exc).__name__}: {exc}"))
        plt.close(fig)
        return None

    plt.close(fig)
    return PlotRecord(
        plot_group="qc",
        workflow_type="all",
        metric_name="completion_summary",
        target_label="all",
        png_path=str(png_path),
        pdf_path=str(pdf_path),
        n_points=len(completion_df),
    )



def generate_plots(
    long_df: pd.DataFrame,
    run_manifest: pd.DataFrame,
    issues_df: pd.DataFrame,
    out_dir: Path,
    plot_min_points: int,
    issues: list[IssueRecord],
    verbose: bool,
) -> list[PlotRecord]:
    """Generate a panel of plots from the long summary table."""
    plots_dir = out_dir / "plots"
    plots_dir.mkdir(parents=True, exist_ok=True)
    plot_records: list[PlotRecord] = []

    if long_df.empty:
        issues.append(IssueRecord("plot_skip", str(plots_dir), "No long-form data available for plotting."))
        completion_plot = generate_completion_plot(
            completion_df=build_completion_table(run_manifest, issues_df),
            plots_dir=plots_dir,
            issues=issues,
        )
        if completion_plot is not None:
            plot_records.append(completion_plot)
        return plot_records

    plot_inputs: list[tuple[str, pd.DataFrame, str, str]] = [
        ("raw", long_df, "metric_name", "metric_name"),
        (
            "baseline_adjusted",
            build_baseline_adjusted_table(long_df),
            "metric_name",
            "baseline_adjusted",
        ),
        (
            "efficiency",
            build_efficiency_table(long_df),
            "metric_name",
            "efficiency",
        ),
    ]

    for plot_group, group_df, _, label_mode in plot_inputs:
        grouped = group_df.groupby(["workflow_type", "metric_name", "target_label"], dropna=False)
        for (workflow_type, metric_name, target_label), plot_df in grouped:
            plot_df = plot_df.dropna(subset=["metric_value", "total_spike_n"])
            if len(plot_df) < plot_min_points:
                issues.append(
                    IssueRecord(
                        "plot_skip",
                        f"{plot_group}:{workflow_type}:{metric_name}:{target_label}",
                        f"Insufficient rows for plotting: {len(plot_df)} < {plot_min_points}.",
                    )
                )
                continue

            if plot_group == "raw":
                y_label = str(metric_name).replace("_", " ")
                title_suffix = str(metric_name)
            elif plot_group == "baseline_adjusted":
                y_label = f"{metric_name.replace('_', ' ')} above baseline"
                title_suffix = f"{metric_name} baseline-adjusted"
            else:
                y_label = f"{metric_name.replace('_', ' ')} / spiked reads"
                title_suffix = f"{metric_name} detection efficiency"

            plot_record = generate_metric_plot(
                plot_df=plot_df,
                out_dir=plots_dir,
                plot_group=plot_group,
                workflow_type=str(workflow_type),
                metric_name=str(metric_name if label_mode == "metric_name" else label_mode),
                target_label=str(target_label),
                issues=issues,
                y_label=y_label,
                title_suffix=title_suffix,
            )
            if plot_record is not None:
                plot_records.append(plot_record)
                log(
                    (
                        f"[INFO] Plotted {plot_record.plot_group} | "
                        f"{plot_record.workflow_type} | "
                        f"{plot_record.target_label} | {plot_record.metric_name}"
                    ),
                    verbose=verbose,
                )

    completion_plot = generate_completion_plot(
        completion_df=build_completion_table(run_manifest, issues_df),
        plots_dir=plots_dir,
        issues=issues,
    )
    if completion_plot is not None:
        plot_records.append(completion_plot)

    return plot_records



def build_summary_statistics(long_df: pd.DataFrame) -> pd.DataFrame:
    """Generate compact summary statistics for each workflow and metric."""
    if long_df.empty:
        return pd.DataFrame(
            columns=[
                "workflow_type",
                "target_label",
                "metric_name",
                "n_rows",
                "n_runs",
                "min_spike_n",
                "max_spike_n",
                "median_metric_value",
                "max_metric_value",
            ]
        )

    grouped = long_df.groupby(["workflow_type", "target_label", "metric_name"], dropna=False)
    summary_df = grouped.agg(
        n_rows=("metric_value", "size"),
        n_runs=("run_name", "nunique"),
        min_spike_n=("total_spike_n", "min"),
        max_spike_n=("total_spike_n", "max"),
        median_metric_value=("metric_value", "median"),
        max_metric_value=("metric_value", "max"),
    ).reset_index()
    return summary_df



def write_tsv(dataframe: pd.DataFrame, path: Path) -> None:
    """Write a DataFrame as a tab-separated file."""
    dataframe.to_csv(path, sep="\t", index=False)



def style_worksheet(worksheet) -> None:
    """Apply standard styling, freeze panes, and autofilter to a worksheet."""
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = max(len(value) for value in values) if values else 0
        width = min(max(max_length + 2, 10), 40)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width



def dataframe_to_excel_sheet(workbook: Workbook, sheet_name: str, dataframe: pd.DataFrame) -> None:
    """Write a DataFrame to a workbook sheet and apply formatting."""
    worksheet = workbook.create_sheet(title=sheet_name[:31])
    if dataframe.empty:
        worksheet.append(["no_data"])
        style_worksheet(worksheet)
        return

    worksheet.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False, name=None):
        worksheet.append(list(row))

    style_worksheet(worksheet)



def write_excel_workbook(
    out_path: Path,
    run_manifest: pd.DataFrame,
    combined_wide: pd.DataFrame,
    combined_long: pd.DataFrame,
    issues_df: pd.DataFrame,
    plot_manifest_df: pd.DataFrame,
    summary_stats_df: pd.DataFrame,
) -> None:
    """Write a formatted Excel workbook containing all main outputs."""
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    dataframe_to_excel_sheet(workbook, "run_manifest", run_manifest)
    dataframe_to_excel_sheet(workbook, "combined_wide", combined_wide)
    dataframe_to_excel_sheet(workbook, "combined_long", combined_long)
    dataframe_to_excel_sheet(workbook, "issues", issues_df)
    dataframe_to_excel_sheet(workbook, "plot_manifest", plot_manifest_df)
    dataframe_to_excel_sheet(workbook, "summary_stats", summary_stats_df)

    workbook.save(out_path)



def main() -> None:
    """Run the discovery, parsing, collation, plotting, and workbook workflow."""
    args = parse_args()
    out_dir = Path(args.out_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    issues: list[IssueRecord] = []

    summary_paths = discover_summary_files(
        input_dirs=args.input_dirs,
        glob_patterns=args.glob_patterns,
        verbose=args.verbose,
    )

    if not summary_paths:
        issues.append(
            IssueRecord(
                "no_summary_files_found",
                str(out_dir),
                "No recognised summary TSV files were found under the input directories.",
            )
        )

    parsed_tables: list[pd.DataFrame] = []
    for summary_path in summary_paths:
        dataframe = parse_summary_file(
            path=summary_path,
            issues=issues,
            verbose=args.verbose,
        )
        if dataframe is not None and not dataframe.empty:
            parsed_tables.append(dataframe)

    run_manifest = build_run_manifest(dataframes=parsed_tables)

    if parsed_tables:
        combined_wide = pd.concat(parsed_tables, ignore_index=True, sort=False)
    else:
        combined_wide = pd.DataFrame()

    combined_long = wide_to_long(dataframe=combined_wide) if not combined_wide.empty else pd.DataFrame()
    summary_stats_df = build_summary_statistics(long_df=combined_long)

    issues_df = pd.DataFrame.from_records([
        {
            "category": issue.category,
            "path": issue.path,
            "detail": issue.detail,
        }
        for issue in issues
    ])
    if issues_df.empty:
        issues_df = pd.DataFrame(columns=["category", "path", "detail"])

    plot_records = generate_plots(
        long_df=combined_long,
        run_manifest=run_manifest,
        issues_df=issues_df,
        out_dir=out_dir,
        plot_min_points=args.plot_min_points,
        issues=issues,
        verbose=args.verbose,
    )

    issues_df = pd.DataFrame.from_records([
        {
            "category": issue.category,
            "path": issue.path,
            "detail": issue.detail,
        }
        for issue in issues
    ])
    if issues_df.empty:
        issues_df = pd.DataFrame(columns=["category", "path", "detail"])

    plot_manifest_df = pd.DataFrame.from_records([
        {
            "plot_group": record.plot_group,
            "workflow_type": record.workflow_type,
            "metric_name": record.metric_name,
            "target_label": record.target_label,
            "png_path": record.png_path,
            "pdf_path": record.pdf_path,
            "n_points": record.n_points,
        }
        for record in plot_records
    ])
    if plot_manifest_df.empty:
        plot_manifest_df = pd.DataFrame(
            columns=[
                "plot_group",
                "workflow_type",
                "metric_name",
                "target_label",
                "png_path",
                "pdf_path",
                "n_points",
            ]
        )

    write_tsv(run_manifest, out_dir / "run_manifest.tsv")
    write_tsv(combined_wide, out_dir / "combined_wide.tsv")
    write_tsv(combined_long, out_dir / "combined_long.tsv")
    write_tsv(summary_stats_df, out_dir / "summary_statistics.tsv")
    write_tsv(issues_df, out_dir / "missing_or_problematic_files.tsv")
    write_tsv(plot_manifest_df, out_dir / "plot_manifest.tsv")

    write_excel_workbook(
        out_path=out_dir / "spikein_summary.xlsx",
        run_manifest=run_manifest,
        combined_wide=combined_wide,
        combined_long=combined_long,
        issues_df=issues_df,
        plot_manifest_df=plot_manifest_df,
        summary_stats_df=summary_stats_df,
    )

    print(f"[INFO] Summary complete. Output directory: {out_dir}")
    print(f"[INFO] Run manifest: {out_dir / 'run_manifest.tsv'}")
    print(f"[INFO] Combined wide table: {out_dir / 'combined_wide.tsv'}")
    print(f"[INFO] Combined long table: {out_dir / 'combined_long.tsv'}")
    print(f"[INFO] Summary statistics: {out_dir / 'summary_statistics.tsv'}")
    print(f"[INFO] Issues table: {out_dir / 'missing_or_problematic_files.tsv'}")
    print(f"[INFO] Plot manifest: {out_dir / 'plot_manifest.tsv'}")
    print(f"[INFO] Excel workbook: {out_dir / 'spikein_summary.xlsx'}")
    print(f"[INFO] N summary files found: {len(summary_paths)}")
    print(f"[INFO] N parsed tables: {len(parsed_tables)}")
    print(f"[INFO] N generated plots: {len(plot_records)}")


if __name__ == "__main__":
    main()
